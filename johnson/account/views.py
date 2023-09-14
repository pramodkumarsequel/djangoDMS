# from django.contrib.auth.models import User
import csv
import uuid
from matplotlib.style import context
from django.db.models import Q
import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import reverse
from django.template.loader import render_to_string
from requests import Response
from rest_framework.response import Response
# from rest_framework import serializers, generics
from rest_framework.views import APIView
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from .decorators import unauthenitcated_user
from .forms import (TaxForm, customer_search_form, CustomerForm, ZoneForm, Zone,
                    DistributorForm, DBPriceMasterForm, ItemForm, HierarchyForm1,
                    HierarchyForm2, HierarchyForm3, HierarchyForm4, HierarchyForm5,
                    SalesForm, DetailSaleForm, UploadFileForm, SaleReturnForm, PurchaseDocumentForm,
                    PurchaseChildForm, PurchaseReturnDocumentForm, DetailOfPurchaseReturnForm, DeliveryNoteForm,
                    DeliveryNoteDetailForm, ReceiptNoteForm, ReceiptNoteDetailForm, distributor_search_form,
                    dbprice_master_searchform, role_search_form, zone_search_form, item_search_form,
                    hierarchy1_search_form, hierarchy2_search_form, hierarchy3_search_form, hierarchy4_search_form,
                    hierarchy5_search_form, user_search_form, Password_Policy_Form, StockForm,
                    LineItemFormset, PurchaseLineItemFormset, SaleReturnLineItem,UserForm,
                    PurchaseReturnLineItem,tax_search_form,UomForm,region_search_form,RegionForm,
                    StateForm,state_search_form,city_search_form, CityForm, warehouseForm,warehouselineItemformset,warehouseItemForm,
                    purchaseinvoicelineitem,purchaseinoviceForm,saleinvoiceform,saleinvoicelineitemForm,orderGeneratorForm,deliveryNoteLineItem

                    )
from .helper import send_forget_password_mail
from .models import (UOM, RegionMaster, RoleMaster, StateMaster, Stock_Child, TaxMaster, User, Customer, Distributor,
                     DBPriceMaster, Item, Hierarchy1, Hierarchy2,
                     Hierarchy3, Hierarchy4, Hierarchy5, Sales_Detail, PurchaseDocument,
                     DetailOfPurchase, DetailsOfPurchaseReturn,
                     Receipt_Note_Detail, Delivery_Note_Details, Profile, Stock_Child,
                     Sales_Return_Detail, CityMaster, Stock, Stock_Child, Sales, Warehouse,warehouseStock,warehouseItemChild,
                     PurchaseinvoicelineItem,salesInvoice,saleinvoicelineItem, OrderGenerator, SalesReturn, PurchaseInvoice,PurchaseReturnDocument,Delivery_Note, Receipt_Note
                     )

from reportsapp.models import MenuMaster, RoleAssignment,stockstatement,Qtystatement
from openpyxl import Workbook
from django.http import HttpResponse
#from reportsapp.decorators import role_required, perm_required, menu_show_as_per_role, check_perm_required

from django.contrib.auth.models import Group, Permission
from account.global_dopdown import *

import json
from decimal import Decimal
from django.http import Http404

#Does quasi the same things as json.loads from here: https://pypi.org/project/dynamodb-json/
class JSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return json.JSONEncoder.default(self, obj)


##################################### Import Sale Details Data via xls ############################


def import_sale_detail(request):
    if request.method == "GET":
        return render(request, 'myapp/account/salebase.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

            excel_data.append(row_data)

        return render(request, 'account/salebase.html', {"excel_data": excel_data})


# logout views
# user logout method..
def user_logout(request):
    logout(request)
    return redirect('login')


# User login views



@unauthenitcated_user
def Login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            request.session['username'] = user.username
            request.session['id'] = user.id
            request.session['email'] = user.email
            #request.session['role'] = user.role
            if user.is_active:
                login(request, user)
                messages.success(request, "You are successfully logged in.")
                return HttpResponseRedirect(reverse('base1'))
            else:
                return HttpResponse("Your account was inactive.")
        else:
            return HttpResponse("Invalid login details given")
    else:
        return render(request, 'account/login.html')



@login_required(login_url='login')
def password_change(request, token):
    context = {}

    try:
        profile_obj = Profile.objects.filter(forgot_password_token=token)

        if request.method == "POST":
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            user_id = request.POST.get('userid')
            print(user_id)

            if user_id is None:
                messages.error(request, 'No userid found.')
                return redirect(f'/password_change/{token}/')

            if new_password != confirm_password:
                messages.error(request, 'No userid found.')
                return redirect(f'/password_change/{token}/')

            user_obj = User.objects.get(id=user_id)
            user_obj.set_password(new_password)
            user_obj.save()
            return redirect('/login/')

        context = {'user_id': profile_obj.user.id}
    except Exception as e:
        print(e)
    return render(request, 'account/password_change.html', context)

@login_required(login_url='login')
def password_reset(request):
    try:
        if request.method == "POST":
            username = request.POST.get('username')
            if not User.objects.filter(username=username).first() or User.objects.filter(email=username).first():
                messages.success(request, 'Not user found with this username.')
                return redirect('/generate_link/$')

            user_obj = User.objects.get(username=username)
            print(user_obj)
            user_mail = User.objects.filter(email=user_obj)
            print(user_mail)
            token = str(uuid.uuid4())
            # profile_obj = Profile.objects.get(user_id = user_obj)
            # profile_obj.forgot_password_token = token
            # profile_obj.save()
            send_forget_password_mail(user_obj, token)
            messages.success(request,
                             'A link has been sent on your email id. Please click on the link to reset your password.')

    except Exception as e:
        print(e)
    return render(request, 'account/password_generate_link.html')


from .forms import MyForm





from account.forms import MyForm
from django.core.cache import cache
import time

# MENU_KEY = "MenuMaster.all"
# @login_required(login_url='login')
# def home(request):
#     form = MyForm()
#     menus = cache.get(MENU_KEY)
#     if not menus:
#         time.sleep(2)  # simulate a slow query.
#         menus = MenuMaster.objects.order_by("id")
#         cache.set(MENU_KEY, menus)
#     c = {'menus': menus}
#     return render(request, 'account/home.html', {'form': form})





########################ZONE View #########################

def export_region_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['content-Disposition'] = 'attachment; filename="region.csv"';
    writer = csv.writer(response)
    writer.writerow(['ID','Zone', 'Region Name', 'Created On','Updated On'

                     ])

    taxes = RegionMaster.objects.all().values_list('id', 'zone__name','Region_Name','created_on', 'updated_on'
                                                      )
    for tax in taxes:
        writer.writerow(tax)

    return response


def export_state_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['content-Disposition'] = 'attachment; filename="state.csv"';
    writer = csv.writer(response)
    writer.writerow(['ID','Zone', 'Region Name','State Name', 'Created On','Updated On'

                     ])

    taxes = StateMaster.objects.all().values_list('id', 'zone','region','State_Name','created_on', 'updated_on'
                                                      )
    for tax in taxes:
        writer.writerow(tax)

    return response


def export_city_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['content-Disposition'] = 'attachment; filename="city.csv"';
    writer = csv.writer(response)
    writer.writerow(['ID','Zone', 'Region Name','State','City Name', 'Created On','Updated On'

                     ])

    taxes = CityMaster.objects.all().values_list('id', 'zone','region','State','City_Name','created_on', 'updated_on'
                                                      )
    for tax in taxes:
        writer.writerow(tax)

    return response



def load_sale_uom(request):
    item_id = request.GET.get('items')
    items = Item.objects.filter(id=item_id).order_by('sale_uom')
    uom=''
    for itm in items:
        uom = itm.sale_uom
    # return render(request, 'account/salereturn_uom_list.html', {'items': items})
    return HttpResponse(json.dumps({'uom': uom}), content_type="application/json")

def load_base_uom(request):
    item_id = request.GET.get('items')
    items = Item.objects.filter(id=item_id).order_by('base_uom')
    base_uom=''
    for itm in items:
        base_uom = itm.base_uom
    # return render(request, 'account/salereturn_uom_list.html', {'items': items})
    return HttpResponse(json.dumps({'uom': base_uom}), content_type="application/json")



def load_disname(request):
    distributor_id = request.GET.get('Distributors')
    customer_id = request.GET.get('Customers')
    distributors = Distributor.objects.filter(id=distributor_id)
    customers = Customer.objects.filter(id=customer_id)
    res1 = ''
    res = ''
    for dis in distributors:
        res = dis.Distributor_Name
    for cus in customers:
        res1 = cus.Customer_Name

    return HttpResponse(json.dumps({'distributor_name': res, 'customer_name': res1}), content_type="application/json")

def load_customer(request):
    try:
        cust_id = request.GET.get('custid')
        customer_name = Customer.objects.get(id=cust_id)
        return HttpResponse(json.dumps({'customer_name': customer_name.Customer_Name}), content_type="application/json")

    except Exception as e:
        customer_name = None
        return HttpResponse(json.dumps({'customer_name': ''}), content_type="application/json")

        


from .forms import deliveryNoteLineItem
from .forms import ReceiptNoteLineItem

############################################## CSV Export File Data #############################
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models.fields.related import ManyToOneRel,ManyToManyField, ManyToManyRel
from phone_field import PhoneNumber
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.styles.numbers import NumberFormat
from openpyxl.writer.excel import save_virtual_workbook
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from account.forms import ExcelImportForm
import pandas as pd  # Requires the pandas library
from django.http import JsonResponse,HttpResponseServerError


def import_data(request):
    if request.method == 'POST'  and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file)  # Use pandas to read the Excel file
        
        # Get the first 100 records or limit the dataframe to 100 rows
        limited_df = df.head(100)
        
        # Counter for tracking the number of successfully imported records
        imported_count = 0
        
        # Loop through the limited data and save it to the model
        for index, row in limited_df.iterrows():
            try:
                # Assuming you have a model called MyModel with fields 'field1', 'field2', etc.
                my_model = User(
                    username = row['username'],
                    userid = row['userid'],
                    email = row['email'],
                    password = row['password'],
                    Phone = row['Phone'],
                    # Map the Excel columns to the model fields
                )
                my_model.save()
                imported_count += 1
            except Exception as e:
                error_message = f"Error importing record at index {index}: {str(e)}"
                return JsonResponse({'status': 'error', 'message': error_message})

        if imported_count < 100:
            return JsonResponse({'status': 'success'})
        else:
            error_message = f"Only {imported_count} records imported out of 100."
            return JsonResponse({'status': 'error', 'message': error_message})
        
        


            
            

def export_user_excel(request):
    queryset = User.objects.all().exclude(username='password')  # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in User._meta.get_fields() if field.name != 'password' if not isinstance(field, models.ImageField) if not isinstance(field, PhoneField) if not isinstance(field, ManyToManyRel) if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user_report.xlsx'
    wb.save(response)

    return response

def export_customer_excel(request):
    queryset = Customer.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Customer._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_report.xlsx'
    wb.save(response)
    return response

def distributor_import_data(request):
    if request.method == 'POST'  and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file)  # Use pandas to read the Excel file
        
        # Get the first 100 records or limit the dataframe to 100 rows
        limited_df = df.head(100)
        
        # Counter for tracking the number of successfully imported records
        imported_count = 0
        
        # Loop through the limited data and save it to the model
        for index, row in limited_df.iterrows():
            try:
                # Assuming you have a model called MyModel with fields 'field1', 'field2', etc.
                my_model = Distributor(
                    dms = row['Distributor_Name'],
                    userid = row['userid'],
                    email = row['email'],
                    password = row['password'],
                    Phone = row['Phone'],
                    # Map the Excel columns to the model fields
                )
                my_model.save()
                imported_count += 1
            except Exception as e:
                error_message = f"Error importing record at index {index}: {str(e)}"
                return JsonResponse({'status': 'error', 'message': error_message})

        if imported_count < 100:
            return JsonResponse({'status': 'success'})
        else:
            error_message = f"Only {imported_count} records imported out of 100."
            return JsonResponse({'status': 'error', 'message': error_message})
        
        
def export_distributor_excel(request):
    queryset = Distributor.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Distributor._meta.get_fields() if not field.related_model if not isinstance(field, PhoneField) if not isinstance(field, ManyToManyRel) if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=distributor_report.xlsx'
    wb.save(response)

    return response

def export_item_excel(request):
    queryset = Item.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Item._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=item_report.xlsx'
    wb.save(response)
    return response

    

def export_hierarchy1_excel(request):
    queryset = Hierarchy1.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Hierarchy1._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hierarchy1_report.xlsx'
    wb.save(response)
    return response


def export_hierarchy2_excel(request):
    queryset = Hierarchy2.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Hierarchy2._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hierarchy2_report.xlsx'
    wb.save(response)
    return response


def export_hierarchy3_excel(request):
    queryset = Hierarchy3.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Hierarchy3._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hierarchy3_report.xlsx'
    wb.save(response)
    return response


def export_hierarchy4_excel(request):
    queryset = Hierarchy4.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Hierarchy4._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hierarchy4_report.xlsx'
    wb.save(response)
    return response


def export_hierarchy5_excel(request):
    queryset = Hierarchy5.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Hierarchy5._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hierarchy5_report.xlsx'
    wb.save(response)
    return response



def export_dbpricemaster_excel(request):
    queryset = DBPriceMaster.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in DBPriceMaster._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dbprice_report.xlsx'
    wb.save(response)
    return response




def export_role_excel(request):
    queryset = RoleMaster.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in RoleMaster._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dbprice_report.xlsx'
    wb.save(response)
    return response




def export_zone_excel(request):
    queryset = Zone.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in Zone._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dbprice_report.xlsx'
    wb.save(response)
    return response




######################### All Details Sale Views Method ################################



def load_Itemname(request):
    try:
        item_id = request.GET.get('items')
        p_order_id=request.GET.get('purchase_order_id')
        dms = DetailOfPurchase.objects.filter(entity_id=p_order_id)
        for d in dms:
            iname=d.item_name
            iqyn = d.Purchases_Quantity
            irate=d.Purchase_Rate
            idis = d.Purchase_Product_Discount
            iserl = d.Purchases_Serial_No
            itotal = d.Total_Amount
            uom=d.purchase_uom
            context = {
                'item_name':iname,
                'uom':uom,
                'iqyn':iqyn,
                'irate':irate,
                'idis':idis,
                'iserl':iserl,
                'itotal':itotal        
            }
            return HttpResponse(json.dumps(context, cls=JSONEncoder), content_type="application/json")
    except ValueError:
        dms=None
        context = {
            '':''
        } 
        return HttpResponse(json.dumps(context, cls=JSONEncoder), content_type="application/json")   


def load_Sales_Item_Name(request):
    item_id = request.GET.get('items')
    if not item_id:
        return HttpResponse(json.dumps({'item_name': None, 'um':None}), content_type="application/json")
    else:
        items = Item.objects.filter(id=item_id)
        for itm in items:
            res = itm.Item_Name
            um=itm.sale_uom.unit_name
            price=itm.MRP
            return HttpResponse(json.dumps({'item_name': res, 'um':um, 'pr':price},cls=JSONEncoder), content_type="application/json")


############################### Customer Search function ################################


def search_customer(request):
    """ search function  """
    if request.method == "POST":
        query_name = request.POST.get('Field_name', None)
        val = request.POST.get('Value' or None)

        if query_name and val:
            results = Customer.objects.filter(Customer__contains=val)
            return HttpResponse({"results": results})

    return redirect('home')



@login_required(login_url='login')
def password_policy(request):
    form = Password_Policy_Form()
    if request.method == "POST":
        form = Password_Policy_Form(request.POST or None)
        minl = request.POST.get('MinChar')
        maxl = request.POST.get('MaxChar')
        passtype = request.POST.get('PassType')
        Passexpdays = request.POST.get('Passexpdays')
        PassexpMsgDays = request.POST.get('PassexpMsgDays')
        Autounlockhour = request.POST.get('Autounlockhour')
        minPassAttempt = request.POST.get('minPassAttempt')
        result = Entity.objects.create(MinChar=minl, MaxChar=maxl, PassType=passtype, Passexpdays=Passexpdays,
                                       PassexpMsgDays=PassexpMsgDays, Autounlockhour=Autounlockhour,
                                       minPassAttempt=minPassAttempt)
        result.save()

    form = Password_Policy_Form()
    return render(request, 'account/password_policy.html', {'form': form})


############################################ Stock Management System#####################################

@login_required(login_url='login')
def stocklist(request):
    form = StockForm()
    stock = Stock_Child.objects.all().values(
        'id',
        'Quantity',
        'Item_Code__TaxRate',
        'Item_Code__Item_Name',
        'Item_Code__sales_detail',
        'Item_Code__sales_detail__Sales_Quantity',
        'Item_Code__sales_return_detail',
        'Item_Code__sales_return_detail__Sales_Quantity',
        'Item_Code__detailofpurchase__Purchases_Quantity',
        'Item_Code__detailsofpurchasereturn__Purchases_Return_Quantity',
        'Item_Code__Item_Code',
        'Item_Code__Hierarchy1__Name',
        'Item_Code__Hierarchy2__Name',
        'Item_Code__Hierarchy3__Name',
        'Item_Code__Hier4__Name',
        'Item_Code__Hier5__Name'
    )
    return render(request, 'account/stockbase.html', {'form': form, 'stock': stock})


######################################### Sales Return Detail View ###############################


from django.http import HttpResponseBadRequest
from django.template import RequestContext
import django_excel as excel


def upload_detailsale(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['file']
            return excel.make_response(filehandle.get_sheet(), "csv")
        else:
            return HttpResponseBadRequest()
    else:
        form = UploadFileForm()
    return render('upload_form.html',
                  {'form': form},
                  context_instance=RequestContext(request))


import json


def load_disname(request):
    distributor_id = request.GET.get('Distributors', '')
    # customer_id = request.GET.get('Customers','')
    print(distributor_id)

    distributors = Distributor.objects.filter(id=distributor_id)
    # customers = Customer.objects.filter(id=customer_id)
    res1 = ''
    res = ''
    for dis in distributors:
        res = dis.Distributor_Name
    # for cus in customers:
    #     res1 = cus.Customer_Name
    print(res)
    # print(res1)
    return HttpResponse(json.dumps({'distributor_name': res}), content_type="application/json")



def load_regions(request):
    try:
        zone_id = request.GET.get('zone')    
        regions = RegionMaster.objects.filter(zone_id=zone_id).order_by('Region_Name')
        states = StateMaster.objects.filter(zone_id=zone_id).order_by('State_Name')
        context = {'regions': regions,'states':states}
        return render(request, 'account/region_dropdown.html', context) 
    except:
        context = {'regions': '','states':''}
        return render(request, 'account/region_dropdown.html', context)

def load_distributor_regions(request):
    try:
        zone_id = request.GET.get('zones')    
        regions = RegionMaster.objects.filter(zone_id=zone_id).order_by('Region_Name')
        context = {'regions': regions}
        return render(request, 'account/region_dropdown.html', context)   
    except:
        context = {'regions': ''}
        return render(request, 'account/region_dropdown.html', context)   

         

# def load_states(request):
#     region_id = request.GET.get('region')    
#     states = StateMaster.objects.filter(region_id=region_id).order_by('State_Name')
#     context = {'states': states}
#     return render(request, 'account/state_dropdown.html', context)    

def load_states(request):
    try:
        zone_id = request.GET.get('zone',None)  
        states = StateMaster.objects.filter(zone_id=zone_id).order_by('State_Name')
        context = {'states': states}
        return render(request, 'account/state_dropdown.html', context)   
    except:
        context = {'states': ''}
        return render(request, 'account/state_dropdown.html', context)
        

def load_cities(request):
    try:
        state_id = request.GET.get('State',None)
        cities = CityMaster.objects.filter(State_id=state_id).order_by('City_Name')
        return render(request, 'account/city_dropdown_list.html', {'cities': cities})   
    except:
        return render(request, 'account/city_dropdown_list.html', {'cities': ''})
        

def dependantfield(request):
    zoneid = request.GET.get('zone', None)
    regionid = request.GET.get('region', None)
    stateid = request.GET.get('state', None)
    cityid = request.GET.get('city', None)
    zone = None
    region = None
    state = None
    if zoneid:
        getzone = Zone.objects.get(id=zoneid)
        region = RegionMaster.objects.filter(zone=getzone)
        
    if regionid:
        getregion = RegionMaster.objects.get(id=regionid)    
        state = StateMaster.objects.filter(region=getregion)
    if stateid:
        getstate = StateMaster.objects.get(id=stateid)
        city = CityMaster.objects.filter(State=getstate)   
    if cityid:
        getcity = CityMaster.objects.get(id=cityid)                  
    zone = Zone.objects.all()
    return render(request,'account/depandentfields.html', locals())    




def export_tax_excel(request):
    queryset = TaxMaster.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in TaxMaster._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tax_report.xlsx'
    wb.save(response)
    return response


def export_uom_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['content-Disposition'] = 'attachment; filename="uom.csv"';
    writer = csv.writer(response)
    writer.writerow(['ID', 'Unit Notation', 'Unit Name', 'Conversion Factor','Status'

                     ])

    taxes = UOM.objects.all().values_list('id', 'unit_notation','unit_name','conversion_factor','status','created_on', 'updated_on'
                                                      )
    for tax in taxes:
        writer.writerow(tax)

    return response


def export_uom_excel(request):
    queryset = UOM.objects.all() # Replace MyModel with your actual model name or queryset

    # Get the fields from the model
    fields = [field for field in UOM._meta.get_fields() if not isinstance(field, ManyToOneRel)]

    # Create a new workbook and get the active sheet
    wb = Workbook()
    sheet = wb.active
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C0C0C0')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write headers to the sheet
    for col_num, field in enumerate(fields, 1):
        cell=sheet.cell(row=1, column=col_num, value=field.verbose_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sheet.auto_filter.ref = sheet.dimensions
    # Write data to the sheet
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field.name)
            sheet.cell(row=row_num, column=col_num, value=str(value))

    # Save the workbook to the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=uom_report.xlsx'
    wb.save(response)
    return response

def load_sale_uom(request):
    item_id = request.GET.get('items')
    items = Item.objects.filter(id=item_id).order_by('sale_uom')
    uom=''
    for itm in items:
        uom = itm.sale_uom
    # return render(request, 'account/salereturn_uom_list.html', {'items': items})
    return HttpResponse(json.dumps({'uom': uom}), content_type="application/json")

def load_base_uom(request):
    item_id = request.GET.get('items')
    items = Item.objects.filter(id=item_id).order_by('base_uom')
    base_uom=''
    for itm in items:
        base_uom = itm.base_uom
    # return render(request, 'account/salereturn_uom_list.html', {'items': items})
    return HttpResponse(json.dumps({'uom': base_uom}), content_type="application/json")



def load_hierachy(request):
    hierarchy1_id = request.GET.get('hierarchy1')    
    hierarchys = Hierarchy2.objects.filter(hierarchy1_id=hierarchy1_id).order_by('Name')
    context = {'hierarchys': hierarchys}
    return render(request, 'account/hierarchy2_dropdown.html', context) 

def load_hierachy2(request):
    hierarchy2_id = request.GET.get('hierarchy2')    
    hierarchys = Hierarchy3.objects.filter(hierarchy2_id=hierarchy2_id).order_by('Name')
    context = {'hierarchys': hierarchys}
    return render(request, 'account/hierarchy3_dropdown.html', context) 

def load_hierarchy4(request):
    hier3id = request.GET.get('hier3id',None)  
    hierarchys = Hierarchy4.objects.filter(hierarchy3_id=hier3id).order_by('Name')
    context = {'hierarchys': hierarchys}
    return render(request, 'account/hierarchy4_dropdown.html', context) 

def load_hierarchy4_ajax(request):
    hier3id = request.GET.get('hierarchy3',None)  
    hierarchys = Hierarchy4.objects.filter(hierarchy3_id=hier3id).order_by('Name')
    print(hierarchys)
    context = {'hierarchys': hierarchys}
    return render(request, 'account/Hierarchy5/hier4_dropdown.html', context) 
 

def load_hier3(request):
    hier2_id = request.GET.get('hierarchy2', None)    
    hierarchys = Hierarchy3.objects.filter(hierarchy2_id=hier2_id).order_by('Name')
    context = {'hierarchys': hierarchys}
    return render(request, 'account/hier3_dropdown.html', context)   

def load_hier5(request):
    hier4id = request.GET.get('hier4id')
    hier5_obj = Hierarchy5.objects.filter(hierarchy4_id=hier4id).order_by('Name')
    context = {'hier5':hier5_obj}
    return render(request, 'account/Item/hier5_dropdown.html', context)

def profile(request):
    return render(request,'account/profile.html')


# MENU_KEY = "MenuMaster.all"
# @login_required(login_url='login')
# def home(request):
#     form = MyForm()
#     menus = cache.get(MENU_KEY)
#     if not menus:
#         time.sleep(2)  # simulate a slow query.
#         menus = MenuMaster.objects.order_by("id")
#         cache.set(MENU_KEY, menus)
#     c = {'menus': menus}
#     return render(request, 'account/home.html', {'form': form})

from reportsapp.serializers import MenuMasterSerializers,SubMasterSerializers
from django.db.models import Avg, Count, Min, Sum

# @login_required(login_url='login')
# def index(request):
#     form = MyForm()
#     try:
#         menulist = MenuMaster.objects.filter(Pmenu=0).order_by('Dord')
#         menuid = [mid.id for mid in MenuMaster.objects.all()]
#         submenulist = MenuMaster.objects.all().filter(Pmenu__in=menuid).order_by('Dord')
#         mainmenu = MenuMasterSerializers(menulist, many=True)
#         data = mainmenu.data
#         current_user_id = request.user.id
#         current_user_role_id = request.user.role.id
#         current_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=19)
#         zone_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=10)
#         role_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=11)
#         item_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=12)
#         hier1_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=13)
#         hier2_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=14)
#         hier3_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=16)
#         hier4_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=17)
#         hier5_menu_perm = MenuMaster.objects.filter(usrole=current_user_role_id).filter(id=18)
#         dis_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=20)
#         dbprice_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=21)
#         users_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=22)
#         region_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=23)
#         state_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=24)
#         city_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=25)
#         tax_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=26)
#         uom_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=27)
#         itm_menu_perm = MenuMaster.objects.filter(usrole = current_user_role_id).filter(id=28)   
#         request.session['itmPerm'] = itm_menu_perm
#         request.session['uomPerm'] = uom_menu_perm
#         request.session['taxPerm'] = tax_menu_perm
#         request.session['cityPerm'] = city_menu_perm
#         request.session['statePerm'] = state_menu_perm
#         request.session['regionPerm'] = region_menu_perm
#         request.session['userPerm'] = users_menu_perm
#         request.session['dbpricePerm'] = dbprice_menu_perm
#         request.session['disPerm'] = dis_menu_perm
#         request.session['hier1Perm'] = hier1_menu_perm
#         request.session['hier2Perm'] = hier2_menu_perm
#         request.session['hier3Perm'] = hier3_menu_perm
#         request.session['hier4Perm'] = hier4_menu_perm
#         request.session['hier5Perm'] = hier5_menu_perm
#         request.session['itemPerm'] = item_menu_perm
#         request.session['rolePerm'] = role_menu_perm
#         request.session['zonePerm'] = zone_menu_perm
#         request.session['hasPerm'] = current_menu_perm
#         request.session['mainM'] = data
#         submenudata = SubMasterSerializers(submenulist, many=True)
#         subdata = submenudata.data      
#         request.session['submenu'] = subdata
#         return render(request, 'account/base.html',{})  
    
#     except Exception as identifier:         
#         return render(request, 'account/base.html',{})  



# def index1(request):
#     items = Item.objects.values('Item_Code').distinct().count()
#     customers = Customer.objects.values('Customer_Name').distinct().count()
#     qs = Sales_Detail.objects.aggregate(Sum('Total_Amount'))
#     items_MRP = Item.objects.aggregate(Sum('MRP'))
#     # ob = DistributorModel.objects.values('DistributorName').distinct().count()
#     # partstate = DistributorModel.objects.values('State').annotate(Sum('Amount'))[:3]
#     # res = DistributorModel.objects.values('DistributorName').annotate(Count('DocID')).order_by().filter(DocID__gt=1)
#     # res1 = DistributorModel.objects.filter(DistributorName__in=[item['DistributorName'] for item in res])
#     #
#     # queryset = DistributorModel.objects.order_by('-DistributorName')[:5]
#     sales_data = Sales.objects.values('Total_Invoice_Amount').annotate(Count('Total_Inventory_Amount'))
#     dataset = Item.objects.values('Item_Name').annotate(Item_MRP=Sum('MRP')).order_by('Item_Name')
#     # distributor_data = Distributor.objects.values('Distributor_Name', 'Distributor_Code')
#     distributor_data = Sales_Detail.objects.values('Sales_Item_Name').annotate(Count('id'))
#     return_dat = [list(q) for q in distributor_data]
#     salesVal = qs['Total_Amount__sum']
#     itm_mrp = items_MRP['MRP__sum']
#     purchases = PurchaseDocument.objects.all().count()
#     sales = Sales.objects.all().count()
#     labels = 'Sale', 'Items'
#     return render(request, 'account/base1.html',{'customers': customers, 'items': items, 'salesval': salesVal, 'itm_mrp': itm_mrp, 'dataset': dataset,
#                    'dataset4': "distributor_data", "purchases":purchases, "sales":sales,
#                    })   


@login_required(login_url='login')
def userList(request):
    models_fields = User._meta.get_fields()
    field_names = [field.verbose_name for field in list(models_fields) if field.get_internal_type() == 'CharField' if field.name != 'password']
    if request.method == "POST":
        select_field = request.POST.get('filter-user-group', None)
        value = request.POST.get('value',None)
        filter_conditions = {} 
        filter_conditions[select_field] = value 
        try:  
            queryset = User.objects.filter(**filter_conditions)
            context = {
            'users':queryset,
            'field_names':field_names,
            }
            return render(request, 'account/usersmain/user_list.html', context)
        except Exception as e:
            messages.error(request, 'Please Select field Name and enter Value.',e)
            return redirect('/user-list/')
            
    users = User.objects.all()
    context = {
        'users':users,
        'field_names':field_names,
    }
    return render(request, 'account/usersmain/user_list.html', context)



def userCreate(request):
    form = UserForm()
    if request.method =="POST":
        form = UserForm(request.POST, None)
        fname = request.POST.get('first_name')
        lname = request.POST.get('last_name')
        uid = request.POST.get('userid')
        uname = request.POST.get('username')
        uemail = request.POST.get('email')
        urole = request.POST.get('role')
        uactive = request.POST.get('is_active')
        mrole = request.POST.getlist('groups[]')
        unumber = request.POST.get('phone_number')
        upass1 = request.POST.get('password1')
        upass2 = request.POST.get('password2')  
        user=User(
            first_name=fname,
            last_name=lname,
            userid=uid,
            username = uname,
            email=uemail,
            role_id=urole,
            Arole = mrole,
            Phone=unumber, 
            password=upass1, 
            
        )
        user.uactive = True
        user.set_password(user.password)
        user.save()
        messages.success(request,'User created successfully.')
        return redirect('/user-list/')  
        
    form = UserForm()
    roles = RoleMaster.objects.all()
    context = {
        'form':form,
        'roles':roles,
    }
    return render(request, 'account/usersmain/user_create.html', context)


def userUpdate(request, pk):
    user = User.objects.get(id=pk)
    print(user)
    if request.method == "POST":
        fname = request.POST.get('first_name')
        lname = request.POST.get('last_name')
        uid = request.POST.get('userid')
        uname = request.POST.get('username')
        uemail = request.POST.get('email')
        urole = request.POST.get('role')
        arole = request.POST.getlist('groups[]')
        phone = request.POST.get('phone')
        print(arole)
        user.first_name=fname
        user.last_name=lname
        user.username=uname
        user.userid=uid
        user.email=uemail
        user.role_id=urole
        user.Arole=arole
        user.Phone=phone
        if user.is_active == True:
            user.save()
        else:
            user.is_active == False
            user.save()    
        messages.success(request,"User updated Successfully.")  
        return redirect('/user-list/')


 
    form = UserForm(instance=user)
    roles = Group.objects.all()
    
    context = {
        'form':form,
        'roles':roles,
    }  
    return render(request,'account/usersmain/user_update.html',context)
    

def userDelete(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.delete()
    return redirect('/user-list/')
    

def customerList(request):
    models_fields = Customer._meta.get_fields()
    field_names = [field for field in list(models_fields) if field.get_internal_type() == 'CharField'] 
    if request.method == "POST":
        select_field = request.POST.get('filter-retailer', None)
        value = request.POST.get('filter-retailer-name', None)
        filter_conditions = {} 
        filter_conditions[select_field] = value 
        try:
            if select_field and value:
                    queryset = Customer.objects.filter(**filter_conditions)    
                    return render(request,'account/Retailer/retailer_list.html',{'customers':queryset, 'field_names':field_names})
            elif select_field or value:
                messages.error(request, 'Please select field name and value')
                return redirect('/retailer-list/')
            
            queryset = Customer.objects.all()  
            return render(request,'account/Retailer/retailer_list.html',{'customers':queryset, 'field_names':field_names})
                  
        except Exception as e:
            messages.success(request, f'{e}')
            return redirect('/retailer-list/')
        
    customers=Customer.objects.none()
    context = {
        'field_names':field_names,
        'customers':customers
    }
    return render(request, 'account/Retailer/retailer_list.html',context)

@login_required(login_url='login')
def customerCreate(request):
    if request.method == "GET":
        form = CustomerForm(request.GET or None)
    elif request.method == 'POST':
        form = CustomerForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, "Retailer Created successfully.")
            return redirect('/customer-list/')
    
    context = {
        'form':form
    }
    return render(request,'account/Retailer/retailer_create.html',context)
    
# @unauthenitcated_user
# @login_required(login_url='login')
def customerUpdate(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Retailer updated successfully.")
            return redirect('/retailer-list/')
            
            
    else: 
        form = CustomerForm(instance=customer)        
        return render(request, 'account/Retailer/retailer_update.html',{'form':form})


def customerDelete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    customer.delete()
    return redirect('/retailer-list/')    


def distributorList(request):
    models_fields = Distributor._meta.get_fields()
    field_names = [field for field in list(models_fields) if field.get_internal_type() == 'CharField']
    if request.method =="POST":
        selected_fields = request.POST.get('selected_fields', '')
        selected_value = request.POST.get('selected_value', '')
        filter_conditions = {} 
        filter_conditions[selected_fields] = selected_value 
        try:
            if selected_fields and selected_value:
                queryset = Distributor.objects.filter(**filter_conditions)
                if not queryset:
                    messages.error(request, 'Please Enter valid Field Value.')
                    return redirect('/distributor-list/')
                
                return render(request,'account/Distributor/distributor_list.html',{'field_names':field_names,'distributors':queryset})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please Select Field and Enter Valid Value.')
                return redirect('/distributor-list/')
                
            
            queryset = Distributor.objects.all()
            return render(request,'account/Distributor/distributor_list.html',{'field_names':field_names,'distributors':queryset})
            
        except Exception as e:
            messages.success(request, 'Please Enter valid Field Value.')
            return redirect('/distributor-list/')
    distributors = Distributor.objects.none()
    context = {
        'distributors':distributors,
        'field_names':field_names,
    }
    return render(request,'account/Distributor/distributor_list.html', context)  


def distributorCreate(request):
    if request.method == "GET":
        form = DistributorForm(request.GET or None)
    elif request.method =="POST":
        form = DistributorForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'Distributor created successfully.') 
            return redirect('/distributor-list/')     
    context = {
        'form':form,
    }
    return render(request, 'account/Distributor/distributor_create.html', context)        
    
    
   
def distributorUpdate(request, pk):
    distributor = get_object_or_404(Distributor, pk=pk)
    if request.method == "POST":
        form = DistributorForm(request.POST , instance=distributor)
        if form.is_valid():
            form.save()
            messages.success(request,'Distributor updated successfully.')
            return redirect('/distributor-list/')
    form = DistributorForm(instance=distributor)
    context = {
        'form':form
    }
    return render(request, 'account/Distributor/distributor_update.html', context)       

def distributorDelete(request, pk):
    distributor = get_object_or_404(Distributor,pk=pk)
    distributor.delete()
    return redirect('/distributor-list/')

 
def itemList(request):
    models_field = Item._meta.get_fields()
    field_names = [field for field in list(models_field) if field.get_internal_type() == "CharField"]
    if request.method=="POST":
        selected_fields = request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Item.objects.filter(**filter_condition)
                return render(request,'account/Item/item_list.html',{'items':queryset,'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and value.')
                return redirect('/item-list/')
            
            queryset = Item.objects.all()
            return render(request,'account/Item/item_list.html',{'items':queryset,'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/item-list/')
            
    items = Item.objects.none()
    context = {
        'items':items,
        'field_names':field_names,
    }
    return render(request,'account/Item/item_list.html', context)


def itemCreate(request):
    if request.method == "GET":
        form = ItemForm(request.GET or None)
    elif request.method == "POST":
        form = ItemForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Item created successfully.')
            return redirect('/item-list/')
    context = {
        'form':form,
    }
    return render(request, 'account/Item/item_create.html', context)    

def itemUpdate(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request,'Item updated successfully.')
            return redirect('/item-list/')
    form = ItemForm(instance=item)
    context = {
        'form':form,
    }        
    return render(request, 'account/Item/item_update.html',context)        

def itemDelete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    item.delete()
    return redirect('/item-list/')


def hierarchy1List(request):
    model_fields = Hierarchy1._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields = request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Hierarchy1.objects.filter(**filter_condition)
                return render(request, 'account/Hierarchy1/hierarchy1_list.html',{'hierarchys':queryset,'field_names':field_names})

            elif selected_fields or selected_value:
                messages.error(request, 'Please enter field name and value')
                return redirect('/hierarchy1-list/')
                
            queryset = Hierarchy1.objects.all()
            return render(request, 'account/Hierarchy1/hierarchy1_list.html',{'hierarchys':queryset,'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/hierarchy1-list/')
        
    hierarchys = Hierarchy1.objects.none()
    context = {
        'hierarchys':hierarchys,
        'field_names':field_names,
    }
    return render(request,'account/Hierarchy1/hierarchy1_list.html', context)

def hierarchy1Create(request):
    if request.method == "GET":
        form = HierarchyForm1(request.GET or None)
    elif request.method == "POST":
        form = HierarchyForm1(request.POST, None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy1 created successfully.')
            return redirect('/hierarchy1-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Hierarchy1/hierarchy1_create.html',context)  

def hierarchy1Update(request,pk):
    hierarchy1 = get_object_or_404(Hierarchy1, pk=pk)
    if request.method == "POST":
        form = HierarchyForm1(request.POST , instance=hierarchy1)
        if form.is_valid():
            form.save()
            messages.success(request,'Hierarchy1 updated successfully.')
            return redirect('/hierarchy1-list/')
    form = HierarchyForm1(instance=hierarchy1)
    context = {
        'form':form,
    }    
    return render(request,'account/Hierarchy1/hierarchy1_update.html', context) 

def hierarchy1Delete(request, pk):
    hierarchy =  get_object_or_404(Hierarchy1, pk=pk)
    hierarchy.delete()
    return redirect('/hierarchy1-list/')    



def hierarchy2List(request):
    model_fields = Hierarchy2._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields = request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Hierarchy2.objects.filter(**filter_condition)
                return render(request,'account/Hierarchy2/hierarchy2_list.html',{'hierarchys2':queryset,'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and value')
                return redirect('/hierarchy2-list/')
            
            queryset = Hierarchy2.objects.all()
            return render(request,'account/Hierarchy2/hierarchy2_list.html',{'hierarchys2':queryset,'field_names':field_names})

        
        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/hierarchy2-list/')
        
    hierarchys2 = Hierarchy2.objects.none()
    context = {
        'hierarchys2':hierarchys2,
        'field_names':field_names,
    }
    return render(request,'account/Hierarchy2/hierarchy2_list.html', context)

def hierarchy2Create(request):
    if request.method == "GET":
        form = HierarchyForm2(request.GET or None)
    elif request.method == "POST":
        form = HierarchyForm2(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'Hierarchy2 created successfully.')
            return redirect('/hierarchy2-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Hierarchy2/hierarchy2_create.html', context)    

def hierarchy2Update(request, pk):
    hierarchy2 = get_object_or_404(Hierarchy2, pk=pk)
    if request.method == "POST":
        form = HierarchyForm2(request.POST, instance=hierarchy2)
        if form.is_valid():
            form.save()
            messages.success(request,'Hierarchy2 updated successfully.')
            return redirect('/hierarchy2-list/')
    form = HierarchyForm2(instance=hierarchy2)
    context = {
        'form':form
    }
    return render(request,'account/Hierarchy2/hierarchy2_update.html', context) 

def hierarchy2Delete(request, pk):
    hierarchy2 = get_object_or_404(Hierarchy2, pk=pk)
    hierarchy2.delete()
    return redirect('/hierarchy2-list/')  


def hierarchy3List(request):
    model_fields = Hierarchy3._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Hierarchy3.objects.filter(**filter_condition)
                return render(request,'account/Hierarchy3/hierarchy3_list.html',{'hierarchys3':queryset, 'field_names':field_names})

            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and value')
                return redirect('/hierarchy3-list/')
            
            queryset = Hierarchy3.objects.all()
            return render(request,'account/Hierarchy3/hierarchy3_list.html',{'hierarchys3':queryset, 'field_names':field_names})
        
        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/hierarchy3-list/')

    hierarchys3=Hierarchy3.objects.none()
    context = {
        'hierarchys3':hierarchys3,
        'field_names':field_names,
    }
    return render(request,'account/Hierarchy3/hierarchy3_list.html',context)

def hierarchy3Create(request):
    if request.method == "GET":
        form = HierarchyForm3(request.GET or None)      
    elif request.method == "POST":
        form = HierarchyForm3(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy3 created successfully.')
            return redirect('/hierarchy3-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Hierarchy3/hierarchy3_create.html',context)

def hierarchy3Update(request, pk):
    hierarchy3 = get_object_or_404(Hierarchy3, pk=pk)
    if request.method == "POST":
        form = HierarchyForm3(request.POST, instance=hierarchy3)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy3 updated successfully.')
            return redirect('/hierarchy3-list/')
    form = HierarchyForm3(instance=hierarchy3)
    context = {
        'form':form
    }
    return render(request, 'account/Hierarchy3/hierarchy3_update.html', context)

def hierarchy3Delete(request, pk):
    hierarchy3 = get_object_or_404(Hierarchy3, pk=pk)
    hierarchy3.delete()
    return redirect('/hierarchy3-list/')               




def hierarchy4List(request):
    model_fields = Hierarchy4._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Hierarchy4.objects.filter(**filter_condition)
                return render(request,'account/Hierarchy4/hierarchy4_list.html', {'hierarchys4':queryset,'field_names':field_names})

            elif selected_fields or selected_value:
                messages.error(request, 'Please enter field name and value')
                return redirect('/hierarchy4-list/')

            queryset = Hierarchy4.objects.all()
            return render(request,'account/Hierarchy4/hierarchy4_list.html', {'hierarchys4':queryset,'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/hierarchy4-list/')
            
    hierarchys4=Hierarchy4.objects.none()
    context = {
        'hierarchys4':hierarchys4,
        'field_names':field_names,
    }
    return render(request,'account/Hierarchy4/hierarchy4_list.html',context)

def hierarchy4Create(request):
    if request.method =="GET":
        form = HierarchyForm4(request.POST or None)
    elif request.method == "POST":
        form = HierarchyForm4(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy4 created successfully.')
            return redirect('/hierarchy4-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Hierarchy4/hierarchy4_create.html',context)

def hierarchy4Update(request, pk):
    hierarchy4 = get_object_or_404(Hierarchy4, pk=pk)
    if request.method == "POST":
        form = HierarchyForm4(request.POST, instance=hierarchy4)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy4 updated successfully.')
            return redirect('/hierarchy4-list/')
    form = HierarchyForm4(instance=hierarchy4)
    context = {
        'form':form
    }
    return render(request, 'account/Hierarchy4/hierarchy4_update.html', context)

def hierarchy4Delete(request, pk):
    hierarchy4 = get_object_or_404(Hierarchy4, pk=pk)
    #hierarchy4.delete()
    return redirect('/hierarchy4-list/')               


def hierarchy5List(request):
    model_fields = Hierarchy5._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields = request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}    
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = Hierarchy5.objects.filter(**filter_condition)
                return render(request,'account/Hierarchy5/hierarchy5_list.html',{'hierarchys5':queryset,'field_names':field_names})
            
            elif selected_fields or selected_value:    
                messages.error(request, 'Please select field and enter value')
                return redirect('/hierarchy5-list/')
            
            queryset = Hierarchy5.objects.all()
            return render(request,'account/Hierarchy5/hierarchy5_list.html',{'hierarchys5':queryset,'field_names':field_names})

        
        except Exception as e:
            messages.error(request, f'{e}') 
            return redirect('/hierarchy5-list/')           
            
    hierarchys5=Hierarchy5.objects.none()
    context = {
        'hierarchys5':hierarchys5,
        'field_names':field_names,
    }
    return render(request,'account/Hierarchy5/hierarchy5_list.html',context)

def hierarchy5Create(request):
    if request.method == "GET":
        form = HierarchyForm5(request.GET or None)
    elif request.method == "POST":
        form = HierarchyForm5(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy5 created successfully.')
            return redirect('/hierarchy5-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Hierarchy5/hierarchy5_create.html',context)

def hierarchy5Update(request, pk):
    hierarchy5 = get_object_or_404(Hierarchy5, pk=pk)
    if request.method == "POST":
        form = HierarchyForm5(request.POST, instance=hierarchy5)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hierarchy5 updated successfully.')
            return redirect('/hierarchy5-list/')
    form = HierarchyForm5(instance=hierarchy5)
    context = {
        'form':form
    }
    return render(request, 'account/Hierarchy5/hierarchy5_update.html', context)

def hierarchy5Delete(request, pk):
    hierarchy5 = get_object_or_404(Hierarchy5, pk=pk)
    hierarchy5.delete()
    return redirect('/hierarchy5-list/') 



def zoneList(request):
    models_fields = Zone._meta.get_fields()
    field_names = [field for field in list(models_fields) if field.get_internal_type() == 'CharField' if not isinstance(field, models.BooleanField)]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
        
        filter_conditions = {} 
        filter_conditions[selected_fields] = selected_value 
        
        try:
            if selected_fields and selected_value:
                queryset = Zone.objects.filter(**filter_conditions)
                return render(request,'account/Zone/zone_list.html',{'zones':queryset, 'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select valid data.')
                return redirect('/zone-list/')
            
            queryset = Zone.objects.all()
            return render(request,'account/Zone/zone_list.html',{'zones':queryset, 'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/zone-list/')
  
    zones=Zone.objects.none()
    context = {
        'zones':zones,
        'field_names':field_names,
    }
    return render(request,'account/Zone/zone_list.html',context)

def zoneCreate(request):
    if request.method == "GET":
        form = ZoneForm(request.GET or None)
    elif request.method == "POST":
        form = ZoneForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zone created successfully.')
            return redirect('/zone-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Zone/zone_create.html',context)

def zoneUpdate(request, pk):
    zone = get_object_or_404(Zone, pk=pk)
    if request.method == "POST":
        form = ZoneForm(request.POST, instance=zone)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zone updated successfully.')
            return redirect('/zone-list/')
    form = ZoneForm(instance=zone)
    context = {
        'form':form
    }
    return render(request, 'account/Zone/zone_update.html', context)

def zoneDelete(request, pk):
    zone = get_object_or_404(Zone, pk=pk)
    zone.delete()
    return redirect('/zone-list/') 





def regionList(request):
    models_field = RegionMaster._meta.get_fields()
    field_names=[field for field in list(models_field) if field.get_internal_type() == 'CharField' if not isinstance(field, models.BooleanField)]
    if request.method == "POST":
        selected_fields = request.POST.get('selected_fields')
        selected_value=request.POST.get('selected_value')
        
        filter_conditions = {} 
        filter_conditions[selected_fields] = selected_value 
        try:
            if selected_value and selected_value:
                queryset = RegionMaster.objects.filter(**filter_conditions)
                return render(request,'account/Regions/region_list.html',{'regions':queryset,'field_names':field_names})
            
            elif selected_value or selected_fields:
                messages.error(request, 'Please  Select Field and  Enter Valid Value.')
                return redirect('/region-list/')
            
            queryset = RegionMaster.objects.all() 
            return render(request,'account/Regions/region_list.html',{'regions':queryset,'field_names':field_names})
        
        except Exception as e:      
            messages.error(request, 'Please Enter Valid.')
            return redirect('/region-list/')
            
    regions=RegionMaster.objects.none()
    context = {
        'regions':regions,
        'field_names':field_names,
    }
    return render(request,'account/Regions/region_list.html',context)

def regionCreate(request):
    if request.method == "GET":
        form = RegionForm(request.GET or None)
    elif request.method == "POST":
        form = RegionForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Region created successfully.')
            return redirect('/region-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Regions/region_create.html',context)

def regionUpdate(request, pk):
    region = get_object_or_404(RegionMaster, pk=pk)
    if request.method == "POST":
        form = RegionForm(request.POST, instance=region)
        if form.is_valid():
            form.save()
            messages.success(request, 'Region updated successfully.')
            return redirect('/region-list/')
    form = RegionForm(instance=region)
    context = {
        'form':form
    }
    return render(request, 'account/Regions/region_update.html', context)

def regionDelete(request, pk):
    region = get_object_or_404(RegionMaster, pk=pk)
    region.delete()
    return redirect('/region-list/') 



def stateList(request):
    model_fields = StateMaster._meta.get_fields()
    field_names = [field for field in list(model_fields) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        
        try:
            if selected_fields and selected_value:
                queryset = StateMaster.objects.filter(**filter_condition)
                return render(request,'account/States/state_list.html',{'states':queryset,'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and Value.')
                return redirect('/state-list/')
            
            queryset = StateMaster.objects.all()
            return render(request,'account/States/state_list.html',{'states':queryset,'field_names':field_names})
            
        except Exception as e:
            messages.error(request, f'{e}')    
                
    states=StateMaster.objects.none()
    context = {
        'states':states,
        'field_names':field_names,
    }
    return render(request,'account/States/state_list.html',context)

def stateCreate(request):
    if request.method == "GET":
        form = StateForm(request.GET or None)
    elif request.method == "POST":
        form = StateForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'State created successfully.')
            return redirect('/state-list/')
    context = {
        'form':form,
    }
    return render(request,'account/States/state_create.html',context)

def stateUpdate(request, pk):
    state = get_object_or_404(StateMaster, pk=pk)
    if request.method == "POST":
        form = StateForm(request.POST, instance=state)
        if form.is_valid():
            form.save()
            messages.success(request, 'State updated successfully.')
            return redirect('/state-list/')
    form = StateForm(instance=state)
    context = {
        'form':form
    }
    return render(request, 'account/States/state_update.html', context)

def stateDelete(request, pk):
    state = get_object_or_404(StateMaster, pk=pk)
    state.delete()
    return redirect('/state-list/') 




def cityList(request):
    models_field = CityMaster._meta.get_fields()
    field_names = [field for field in list(models_field) if field.get_internal_type() == 'CharField']
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields')
        selected_value = request.POST.get('selected_value')
       
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        try:
            if selected_fields and selected_value:
                queryset = CityMaster.objects.filter(**filter_condition)
                return render(request,'account/Citys/city_list.html',{'citys':queryset,'field_names':field_names})

            elif selected_fields or selected_value:
                messages.error(request, 'Please Select valid field name and value.')
                return redirect('/city-list/')
            
            queryset = CityMaster.objects.all()
            return render(request,'account/Citys/city_list.html',{'citys':queryset,'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/city-list/')
        
    citys=CityMaster.objects.none()
    context = {
        'citys':citys,
        'field_names':field_names,
    }
    return render(request,'account/Citys/city_list.html',context)

def cityCreate(request):
    if request.method == "GET":
        form = CityForm(request.GET or None)
    elif request.method == "POST":
        form = CityForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request, 'City created successfully.')
            return redirect('/city-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Citys/city_create.html',context)

def cityUpdate(request, pk):
    city = get_object_or_404(CityMaster, pk=pk)
    if request.method == "POST":
        form = CityForm(request.POST, instance=city)
        if form.is_valid():
            form.save()
            messages.success(request, 'City updated successfully.')
            return redirect('/city-list/')
    form = CityForm(instance=city)
    context = {
        'form':form
    }
    return render(request, 'account/Citys/city_update.html', context)

def cityDelete(request, pk):
    city = get_object_or_404(CityMaster, pk=pk)
    city.delete()
    return redirect('/city-list/') 





def uomList(request):
    models_field = UOM._meta.get_fields()
    field_names = [field for field in list(models_field) if field.get_internal_type() == 'CharField']
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        try:
            if selected_fields and selected_value:
                queryset = UOM.objects.filter(**filter_condition)
                return render(request,'account/Uom/uom_list.html',{'uoms':queryset, 'field_names':field_names})
            
            elif selected_value or selected_fields:
                messages.error(request, 'Please select field name and field value.')
                return redirect('/uom-list/')
            
            queryset = UOM.objects.all()
            return render(request,'account/Uom/uom_list.html',{'uoms':queryset, 'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/uom-list/')
        
    uoms=UOM.objects.none()
    context = {
        'uoms':uoms,
        'field_names':field_names,
    }
    return render(request,'account/Uom/uom_list.html',context)


def uomCreate(request):
    if request.method == "GET":
        form = UomForm(request.GET or None)
    elif request.method =="POST":
        form = UomForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'Uom created successfully!.')
            return redirect('/uom-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Uom/uom_create.html', context)        


def uomUpdate(request, pk):
    uom = get_object_or_404(UOM, pk=pk)
    if request.method == "POST":
        form = UomForm(request.POST, instance=uom)
        if form.is_valid():
            form.save()
            messages.success(request, 'Uom updated successfully.')
            return redirect('/uom-list/')
    form = UomForm(instance=uom)
    context = {
        'form':form
    }
    return render(request, 'account/Uom/uom_update.html', context)

def uomDelete(request, pk):
    uom = get_object_or_404(UOM, pk=pk)
    uom.delete()
    return redirect('/uom-list/')
    
def taxList(request):
    models_field = TaxMaster._meta.get_fields()
    field_names = [field for field in list(models_field) if field.get_internal_type() == 'CharField']
    if request.method == "POST":
        selected_fields = request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        try:
            if selected_fields and selected_value:
                queryset = TaxMaster.objects.filter(**filter_condition)
                return render(request,'account/Tax/tax_list.html',{'taxes':queryset, 'field_names':field_names})

            elif selected_fields or selected_value:
                messages.error(request, 'Please enter field name and field value.')
                return redirect('/tax-list/')
            
            queryset = TaxMaster.objects.all()
            return render(request,'account/Tax/tax_list.html',{'taxes':queryset, 'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/tax-list/')
    taxes=TaxMaster.objects.none()
    context = {
        'taxes':taxes,
        'field_names':field_names,
    }
    return render(request,'account/Tax/tax_list.html',context)


def taxCreate(request):
    if request.method == "GET":
        form = TaxForm(request.GET or None)
    elif request.method =="POST":
        form = TaxForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'Tax created successfully!.')
            return redirect('/tax-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Tax/tax_create.html', context)        


def taxUpdate(request, pk):
    tax = get_object_or_404(TaxMaster, pk=pk)
    if request.method == "POST":
        form = TaxForm(request.POST, instance=tax)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tax updated successfully.')
            return redirect('/tax-list/')
    form = TaxForm(instance=tax)
    context = {
        'form':form
    }
    return render(request, 'account/Tax/tax_update.html', context)

def taxDelete(request, pk):
    tax = get_object_or_404(TaxMaster, pk=pk)
    tax.delete()
    return redirect('/tax-list/')
    


def dbpriceList(request):
    models_field = DBPriceMaster._meta.get_fields()
    field_names = [field for field in list(models_field)]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields' or None)
        selected_value = request.POST.get('selected_value' or None)
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        try:
            if selected_fields and selected_value:
                queryset = DBPriceMaster.objects.filter(**filter_condition)
                return render(request,'account/Dbprice/dbprice_list.html',{'dbprices':queryset, 'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and value.')
                return redirect('/dbprice-list/')  
            
            queryset = DBPriceMaster.objects.all()
            return render(request,'account/Dbprice/dbprice_list.html',{'dbprices':queryset, 'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/dbprice-list/')

    dbprices=DBPriceMaster.objects.none()
    context = {
        'dbprices':dbprices,
        'field_names':field_names,
        
    }
    return render(request,'account/Dbprice/dbprice_list.html',context)


def dbpriceCreate(request):
    if request.method == "GET":
        form = DBPriceMasterForm(request.GET or None)
    elif request.method =="POST":
        form = DBPriceMasterForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'DBPrice created successfully!.')
            return redirect('/dbprice-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Dbprice/dbprice_create.html', context)        


def dbpriceUpdate(request, pk):
    dbprice = get_object_or_404(DBPriceMaster, pk=pk)
    if request.method == "POST":
        form = DBPriceMasterForm(request.POST, instance=dbprice)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dbprice updated successfully.')
            return redirect('/dbprice-list/')
    form = DBPriceMasterForm(instance=dbprice)
    context = {
        'form':form
    }
    return render(request, 'account/Dbprice/dbprice_update.html', context)

def dbpriceDelete(request, pk):
    dbprice = get_object_or_404(DBPriceMaster, pk=pk)
    dbprice.delete()
    return redirect('/dbprice-list/')


def warehouseList(request):
    models_field = Warehouse._meta.get_fields()
    field_names = [field for field in list(models_field) if field.get_internal_type() == "CharField"]
    if request.method == "POST":
        selected_fields=request.POST.get('selected_fields')
        selected_value=request.POST.get('selected_value')
        
        filter_condition = {}
        filter_condition[selected_fields] = selected_value
        try:
            if selected_fields and selected_value:
                queryset = Warehouse.objects.filter(**filter_condition)
                return render(request,'account/Warehouse/warehouse_list.html',{'warehouses':queryset, 'field_names':field_names})
            
            elif selected_fields or selected_value:
                messages.error(request, 'Please select field name and field value.')
                return redirect('/warehouse-list/')
            
            queryset = Warehouse.objects.all()
            return render(request,'account/Warehouse/warehouse_list.html',{'warehouses':queryset, 'field_names':field_names})

        except Exception as e:
            messages.error(request, f'{e}')
            return redirect('/warehouse-list/')

    warehouses=Warehouse.objects.none()
    context = {
        'warehouses':warehouses,
        'field_names':field_names,
    }
    return render(request,'account/Warehouse/warehouse_list.html',context)


def warehouseCreate(request):
    if request.method == "GET":
        form = warehouseForm(request.GET or None)
    elif request.method =="POST":
        form = warehouseForm(request.POST or None)
        if form.is_valid():
            form.save()
            messages.success(request,'Warehouse created successfully!.')
            return redirect('/warehouse-list/')
    context = {
        'form':form,
    }
    return render(request,'account/Warehouse/warehouse_create.html', context)        


def warehouseUpdate(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == "POST":
        form = warehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            messages.success(request, 'Warehouse updated successfully.')
            return redirect('/warehouse-list/')
    form = warehouseForm(instance=warehouse)
    context = {
        'form':form
    }
    return render(request, 'account/Warehouse/warehouse_update.html', context)

def warehouseDelete(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    warehouse.delete()
    return redirect('/warehouse-list/')

from django.db import IntegrityError

#@login_required(login_url='login')
def warehouseOpeningStock(request):
    if request.method == "POST":
        parent_form = warehouseItemForm(request.POST or None)
        formset = warehouselineItemformset(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        w_house_object = warehouseItemChild.objects.filter(w_house=parent_instance.warehouseCode.id, Item_Code_id=child_instance.Item_Code.id).exists()
                        if w_house_object:
                            transaction.set_rollback(True) 
                            messages.warning(request,'Already exists!.', extra_tags="alert-warning") 
                            return redirect('/warehouseopeningstock-create/') 
                                
                        child_instance.warehouseID = parent_instance.id
                        child_instance.w = parent_instance
                        child_instance.w_house=parent_instance.warehouseCode.id
                        stockstatement.objects.create(document_type="Opening Balance",
                                                        docid=parent_instance.id,
                                                            doc_date=parent_instance.Date,
                                                            warehouse_id=parent_instance.warehouseCode.id,
                                                            w_house = parent_instance.warehouseCode,
                                                            warehouse_name=parent_instance.warehouseCode.wname,
                                                            items=child_instance.Item_Code.id,
                                                            itm=child_instance.Item_Code,
                                                            item_name=child_instance.Item_Code.Item_Name,
                                                            Qty=child_instance.Quantity,
                                                            uom=child_instance.base_uom,
                                                            amount=child_instance.Total_Amount
                                                            
                                                        )
                        child_instance.save()
                    messages.success(request, f'Warehouse {parent_instance.w_order_no} created successfully!')        
                    return redirect('/warehouseopeningstock-create/')   
                                        
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.error(request, f'An error occurred during form submission. Please try again later.{str(e)}', extra_tags="alert-danger")
                return redirect('/warehouseopeningstock-create/') 
    else:
        parent_form = warehouseItemForm()
        formset = warehouselineItemformset()
    context={
            'form':parent_form,
            'formset':formset,
        }

    return render(request, 'account/Openingbalance/openingbalance_create.html',context)        
            



#popup CGST& SGST AMount 

def get_cgst_sgst_amount(request):
    gst_type_id = request.GET.get('id_gst_type')
    try:
        gstquery = TaxMaster.objects.get(id=gst_type_id)   
        context = {
            'SGST_AMOUNT':gstquery.SGST,
            'CGST_AMOUNT':gstquery.CGST
        } 
        return HttpResponse(json.dumps(context,cls=JSONEncoder), content_type="application/json")

    except ValueError:
        context = {
            'SGST_AMOUNT':None,
            'CGST_AMOUNT':None
        } 
        return HttpResponse(json.dumps(context,cls=JSONEncoder), content_type="application/json")


# populate warehouse Name

def load_warehouseName(request):
    wid = request.GET.get('wh_id' or None)
    if not wid:
        return HttpResponse(json.dumps({'whName':''}), content_type="application/json")
    ids = Warehouse.objects.get(id=wid)
    return HttpResponse(json.dumps({'whName':ids.wname}), content_type="application/json")
 

        


from datetime import datetime  
from django.utils.dateformat import DateFormat
from django.utils.formats import get_format 
 

@login_required(login_url='login')
def purchaselineitemCreate(request):
    if request.method == "POST":
        parent_form = PurchaseDocumentForm(request.POST or None)
        formset = PurchaseLineItemFormset(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        child_instance.purchase = parent_instance
                        child_instance.entity_id = parent_instance.id
                        child_instance.bal_qty = child_instance.Purchases_Quantity
                        child_instance.save()
                messages.success(request, f'Purchase Document {parent_instance.SAP_Order_No} created successfully!')        
                return redirect('/purchase-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/purchase-create/') 
    else:
        parent_form = PurchaseDocumentForm()
        formset = PurchaseLineItemFormset()
    context={
            'form':parent_form,
            'formset':formset,
        }

    return render(request, 'account/Purchase/purchase_create.html', context)        





@login_required(login_url='login')
def receiptynoteCreate(request):
    if request.method == "POST":
        parent_form = ReceiptNoteForm(request.POST or None)
        formset = ReceiptNoteLineItem(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        stock = get_object_or_404(DetailOfPurchase, entity_id=parent_instance.p_order_no.id, items=child_instance.items.id)
                        stock.bal_qty-=child_instance.Quantity
                        child_instance.MRN_QTY = child_instance.Quantity
                        child_instance.entity_id = parent_instance.id
                        child_instance.receipt = parent_instance
                        stockstatement.objects.create(document_type="Material Receipt Note",
                                                docid=parent_instance.id,
                                                doc_date=parent_instance.receipt_note_date,
                                                warehouse_id=parent_instance.warehouse.id,
                                                w_house =parent_instance.warehouse, 
                                                warehouse_name=parent_instance.warehouse.wname,
                                                items=child_instance.items.id,
                                                itm=child_instance.items,
                                                item_name=child_instance.Item_Name,
                                                Qty=child_instance.Quantity,
                                                uom=child_instance.uom,
                                                amount=child_instance.Total_Amount
                                                
                                            )
                        
                        # update stock summary
                        child_instance.save()
                        stock.save()
                messages.success(request, f'Warehouse {parent_instance.recipt_no} created successfully!')        
                return redirect('/receiptnote-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/receiptnote-create/') 
    else:
        parent_form = ReceiptNoteForm()
        formset = ReceiptNoteLineItem()
    context={
            'form':parent_form,
            'formset':formset,
            'stocks': DetailOfPurchase.objects.all(),
        }

    return render(request, 'account/Receiptnote/receiptnote_create.html', context)




@login_required(login_url='login')
def purchaseinvoiceCreate(request):
    if request.method == "POST":
        parent_form = purchaseinoviceForm(request.POST or None)
        formset = purchaseinvoicelineitem(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        child_instance.PI_QTY=child_instance.Quantity
                        stock = get_object_or_404(Receipt_Note_Detail, entity_id=parent_instance.mrn_no.id, items=child_instance.items.id)
                        stock.MRN_QTY -= child_instance.Quantity
                        child_instance.parent_id = parent_instance.id
                        child_instance.pinvoice=parent_instance
                        stock.save()   
                        child_instance.save()
                messages.success(request, f'Purchase Invoice {parent_instance.purchase_invoice_no} created successfully!')        
                return redirect('/purchaseinvoice-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/purchaseinvoice-create/') 
    else:
        parent_form = purchaseinoviceForm()
        formset = purchaseinvoicelineitem()
    context={
            'form':parent_form,
            'formset':formset,
            "stocks": Receipt_Note_Detail.objects.all(),
            "items": Item.objects.all(),
        }

    return render(request, 'account/Purchaseinvoice/purchaseinvoice_create.html', context)     




@login_required(login_url='login')
def purchasereturnlineitemCreate(request):
    if request.method == "POST":
        parent_form = PurchaseReturnDocumentForm(request.POST or None)
        formset = PurchaseReturnLineItem(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        purchase_invoice = PurchaseinvoicelineItem.objects.filter(parent_id=parent_instance.pi_no.id).filter(items=child_instance.items).get()
                        purchase_invoice.PI_QTY -= child_instance.Purchases_Return_Quantity
                        child_instance.entity_id = parent_instance.id
                        child_instance.preturn = parent_instance
                        purchase_invoice.save()
                        stockstatement.objects.create(document_type="Purchase Return",
                                                docid=parent_instance.pi_no.id,
                                                doc_date=parent_instance.Bill_Date,
                                                warehouse_id=parent_instance.warehouse.id,
                                                w_house =parent_instance.warehouse, 
                                                warehouse_name=parent_instance.warehouse.wname,
                                                items=child_instance.items.id,
                                                itm = child_instance.items,
                                                item_name=child_instance.items.Item_Name,
                                                Qty=child_instance.Purchases_Return_Quantity,
                                                uom=child_instance.purchase_uom,
                                                amount=child_instance.Total_Amount
                                                
                                                )

                        child_instance.save()
                messages.success(request, f'Purchase Return {parent_instance.SAP_Order_No} created successfully!')        
                return redirect('/purchase-return-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/purchase-return-create/') 
    else:
        parent_form = PurchaseReturnDocumentForm()
        formset = PurchaseReturnLineItem()
    context={
            'form':parent_form,
            'formset':formset,
            "stocks": PurchaseinvoicelineItem.objects.all(),
        }
    return render(request, 'account/Purchasereturn/purchasereturn_create.html', context) 






@login_required(login_url='login')
def saleCreate(request):
    if request.method == "POST":
        parent_form = SalesForm(request.POST or None)
        formset = LineItemFormset(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        child_instance.sale = parent_instance
                        child_instance.entity_id = parent_instance.id
                        child_instance.bal_qty = child_instance.Sales_Quantity
                        child_instance.save()
                messages.success(request, f'Sales Order {parent_instance.sales_order_no} created successfully!')        
                return redirect('/saleCreate/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/saleCreate/') 
    else:
        parent_form = SalesForm()
        formset = LineItemFormset()

    context={
            'form':parent_form,
            'formset':formset,
            "stocks": warehouseItemChild.objects.all(),
            'items': Item.objects.all()
        }
    return render(request, 'account/SaleDocument/sale_create.html', context)	








def load_so_date(request):
    try:
        soid = request.GET.get('soid')
        so_obj = Sales.objects.get(id=soid)
        so_date = so_obj.Bill_Date
        sdate = DateFormat(so_date)
        s_date = sdate.format('Y-m-d')
        return HttpResponse(json.dumps({'s_date':s_date}), content_type="application/json")
    
    except ValueError:
        soid=None
        s_date=None
        return HttpResponse(json.dumps({'s_date':s_date}), content_type="application/json")
        


@login_required(login_url='login')
def deliverynoteCreate(request):
    if request.method == "POST":
        parent_form = DeliveryNoteForm(request.POST or None)
        formset = deliveryNoteLineItem(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        child_instance.DN_QTY = child_instance.Quantity
                        stock_obj=Sales_Detail.objects.filter(entity_id=parent_instance.sales_order_no.id).filter(Sales_Item_Code=child_instance.items).get()
                        stock_obj.bal_qty -=child_instance.Quantity
                        child_instance.entity_id = parent_instance.id
                        child_instance.note = parent_instance
                        stockstatement.objects.create(document_type="Delivery Note",
                                                docid=parent_instance.id,
                                                doc_date=parent_instance.delivery_note_date,
                                                warehouse_id=parent_instance.warehouse.id,
                                                w_house =parent_instance.warehouse,
                                                warehouse_name=parent_instance.warehouse.wname,
                                                items=child_instance.items.id,
                                                itm = child_instance.items,
                                                item_name=child_instance.items.Item_Name,
                                                Qty=child_instance.Quantity,
                                                uom=child_instance.uom,
                                                amount=child_instance.Amount
                                                )
                            
                        stock_obj.save()
                        child_instance.save()          
                    messages.success(request, f'Delivery Note {parent_instance.delivery_note_no} created successfully!')        
                    return redirect('/deliverynote-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/deliverynote-create/') 
    else:
        parent_form = DeliveryNoteForm()
        formset = deliveryNoteLineItem() 
    context={
            'form':parent_form,
            'formset':formset,
            "stocks": Sales_Detail.objects.all(),
        }
    return render(request, 'account/Deliverynote/deliverynote_create.html', context)	


def load_di_date(request):
    try:
        dlid = request.GET.get('dlid')
        so_obj = Delivery_Note.objects.get(id=dlid)
        di_date = so_obj.delivery_note_date
        didate = DateFormat(di_date)
        d_date = didate.format('Y-m-d')
        return HttpResponse(json.dumps({'d_date':d_date}), content_type="application/json")
    except Exception as esc:
        d_date=None
        return HttpResponse(json.dumps({'d_date':''}), content_type="application/json")
        



@login_required(login_url='login')
def saleseinvoiceCreate(request):
    if request.method == "POST":
        parent_form = saleinvoiceform(request.POST or None)
        formset = saleinvoicelineitemForm(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        child_instance.SI_QTY = child_instance.Quantity
                        stock_obj = Delivery_Note_Details.objects.filter(entity_id=parent_instance.dl_note_no.id).filter(items=child_instance.Item_Code.id).get()
                        stock_obj.DN_QTY -= child_instance.Quantity
                        child_instance.parent_id = parent_instance.id  
                        child_instance.inv = parent_instance                       
                        child_instance.save()
                        stock_obj.save()
                messages.success(request, f'Sales Invoice {parent_instance.Invoice_No} created successfully!')        
                return redirect('/sales-invoice-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/sales-invoice-create/') 
    else:
        parent_form = saleinvoiceform()
        formset = saleinvoicelineitemForm() 
    context={
            'form':parent_form,
            'formset':formset,
            "stocks": Delivery_Note_Details.objects.all(),
        }
    return render(request, 'account/Saleinvoice/saleinvoice_create.html', context)







@login_required(login_url='login')
def salereturnCreate(request):
    if request.method == "POST":
        parent_form = SaleReturnForm(request.POST or None)
        formset = SaleReturnLineItem(request.POST or None)      
        if parent_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    parent_instance = parent_form.save()
                    for form in formset:
                        child_instance = form.save(commit=False)
                        sale_inv_obj = saleinvoicelineItem.objects.filter(parent_id=parent_instance.sale_invoice_no.id, Item_Code=child_instance.Sales_Item_Code.id).get()
                        sale_inv_obj.SI_QTY -=child_instance.Sales_Quantity
                        sale_inv_obj.save()
                        stockstatement.objects.create(document_type="Sales Return",
                                                docid=parent_instance.sale_invoice_no.id,
                                                doc_date=parent_instance.sales_return_date,
                                                warehouse_id=parent_instance.warehouse.id,
                                                w_house =parent_instance.warehouse,
                                                warehouse_name=parent_instance.warehouse.wname,
                                                items=child_instance.Sales_Item_Code.id,
                                                item_name=child_instance.Sales_Item_Name,
                                                itm = child_instance.Sales_Item_Code,
                                                Qty=child_instance.Sales_Quantity,
                                                uom=child_instance.sale_uom,
                                                amount=child_instance.Total_Amount
                                                )
                        child_instance.entity_id = parent_instance.id  
                        child_instance.salereturn = parent_instance                       
                        child_instance.save()
                messages.success(request, f'Sales Return {parent_instance.sales_return_no} created successfully!')        
                return redirect('/salereturn-create/')  
            
            except Exception as e:
                # Handle the exception here
                # Optionally, you can display an error message to the user
                messages.warning(request, f'An error occurred during form submission. Please try again later.{str(e)}')
                return redirect('/salereturn-create/') 
    else:
        parent_form = SaleReturnForm()
        formset = SaleReturnLineItem() 
    context={
            'form':parent_form,
            'formset':formset,
            "stocks": saleinvoicelineItem.objects.all(),
        }
    return render(request, 'account/Salereturn/salesreturn_create.html', context)








def load_si_date(request):
    si_id = request.GET.get('si_id')
    so_obj = salesInvoice.objects.get(id=si_id)
    di_date = so_obj.Bill_Date
    didate = DateFormat(di_date)
    d_date = didate.format('Y-m-d')
    return HttpResponse(json.dumps({'si_id':d_date}), content_type="application/json")





def load_pidate(request):
    pid=request.GET.get('pid')
    pidate_object = PurchaseInvoice.objects.get(id=pid)
    dt = pidate_object.purchase_invoice_date
    sup_inv_no = pidate_object.Supplier_Invoice_No
    sup_inv_date=pidate_object.SAP_Order_Date
    df=DateFormat(dt)
    supdate=DateFormat(sup_inv_date)
    sdate=supdate.format('Y-m-d')
    mrndate = df.format('Y-m-d')
    context = {
        'piDate':mrndate,
        'sup':sup_inv_no,
        'supdate':sdate,
        'dcode':pidate_object.Distributors.Distributor_Code,
        'dname':pidate_object.Distributor_Name,
        'cgst':pidate_object.CGST_AMOUNT,
        'sgst':pidate_object.SGST_AMOUNT,
        'tgst':pidate_object.Total_GST,
        'td':pidate_object.total_discount_amount,
        'tinv':pidate_object.Total_Inventory_Amount,
        'tinvoice':pidate_object.Total_Invoice_Amount,
       
    }
    return HttpResponse(json.dumps(context, cls=JSONEncoder), content_type="application/json")


       
        

from .forms import ReceiptNoteLineItem

#@login_required(login_url='login')

def get_pur_order_id(request):
    pi=request.GET.get('orderid')
    return HttpResponse(json.dumps({'or':pi}, cls=JSONEncoder), content_type="application/json")






#Order Combination generated page 

def get_model(model):
    return model._meta.get_fields('include_hidden=True')

def get_model_fields(request):
    Model = request.GET.get('fields')
    if Model=="Sales":
        fields=[f.verbose_name for f in Sales._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    elif Model == "salesInvoice":
        fields=[f.verbose_name for f in salesInvoice._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif Model == "SalesReturn":
        fields=[f.verbose_name for f in SalesReturn._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif Model == "PurchaseDocument":
        fields=[f.verbose_name for f in PurchaseDocument._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif Model == "PurchaseInvoice":
        fields=[f.verbose_name for f in PurchaseInvoice._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    
    elif Model == "PurchaseReturnDocument":
        fields=[f.verbose_name for f in PurchaseReturnDocument._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif Model == "Delivery_Note":
        fields=[f.verbose_name for f in Delivery_Note._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    
    elif Model == "Receipt_Note":
        fields=[f.verbose_name for f in Receipt_Note._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif Model == "openingbalance":
        fields=[f.verbose_name for f in Warehouse._meta.get_fields() if f.many_to_one]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})


    

def get_fields(request):
    field = request.GET.get('fields')
    if field == "Distributor Code":
        fields=[f.verbose_name for f in Distributor._meta.get_fields() if not f.related_model]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif field == "Customer Code":
        fields=[f.verbose_name for f in Customer._meta.get_fields() if not f.related_model]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
    
    elif field == "GST Type":
        fields=[f.verbose_name for f in TaxMaster._meta.get_fields() if not f.related_model]
        return render(request, 'account/Order/field_dropdown.html', {'fields': fields})
        
    return render(request, 'account/Order/field_dropdown.html', {'fields': field})
    
from django.db.models import Max
def ordergenerateCreate(request):  
    if request.method == "POST":
        doctype = request.POST.get('documentType')
        fld_selection = request.POST.get('Field_selection')
        prefix = request.POST.get('prfx')
        seq = request.POST.get('seqnum')
        doc_field = request.POST.get('fields')
        f_for_prfx = request.POST.get('field')
        sufix = request.POST.get('sufx')
        seperator = request.POST.get('seprator_choice')
        use_FY_as_prfx = request.POST.get('use_FY_as_prfx')
        use_month_as_prfx = request.POST.get('use_month_as_prfx')
        use_month_as_sufx = request.POST.get('use_month_as_sufx')
        use_FY_as_sufx = request.POST.get('use_FY_as_sufx')
        ob = OrderGenerator.objects.filter(documentType=doctype).exists()
        if ob:
            messages.warning(request, 'Already Exists.')
            return redirect('/number-series-generate/')     
        order=OrderGenerator(
            documentType=doctype,
            prfx=prefix,
            Field_selection=fld_selection,
            seqnum=seq,
            Document_for_prefix=doc_field,
            Doc_field=f_for_prfx,
            seprator_choice=seperator    
        )
        order_obj=order
        last_order = OrderGenerator.objects.all().order_by("-pk").first()
        if last_order:
            last_pk = last_order.pk
        order_com = prefix +str(seperator)+ str(last_pk+1).zfill(len(seq))  
        order_obj.order_combination=order_com
        order_obj.save()
        messages.success(request, 'Order Generated sequence created successfully.')
        return redirect('/number-series-generate/')
    intaial_data = {
        'Field_selection':'0'
    }
    form = orderGeneratorForm(initial=intaial_data)
    orderGens = OrderGenerator.objects.all()
    context = {
        'form':form,
        'orderGens':orderGens,
    }
    return render(request,'account/Order/orderbase.html', context)    



def ordergenerateUpdate(request, pk):
    order = get_object_or_404(OrderGenerator,pk=pk)
    if request.method == "POST":
        doctype = request.POST.get('documentType')
        fld_selection = request.POST.get('Field_selection')
        prefix = request.POST.get('prfx')
        sufix = request.POST.get('sufx')
        seq = request.POST.get('seqnum')
        doc_field = request.POST.get('fields')
        f_for_prfx = request.POST.get('field')
        seperator = request.POST.get('seprator_choice')
        use_FY_as_prfx = request.POST.get('use_FY_as_prfx')
        use_month_as_prfx = request.POST.get('use_month_as_prfx')
        use_month_as_sufx = request.POST.get('use_month_as_sufx')
        use_FY_as_sufx = request.POST.get('use_FY_as_sufx')
        print(fld_selection)
        od = str(prefix)+str(seperator) + str(seq) +str(sufix)
        if fld_selection == '0':
            order.documentType=doctype
            order.Field_selection=fld_selection
            order.prfx=prefix
            order.sufx=sufix
            order.seprator_choice=seperator
            order.seqnum=seq
            order.order_combination=od
            order.save()
        else:
            order.documentType=doctype
            order.Field_selection=fld_selection
            order.prfx=prefix
            order.sufx=sufix
            order.seprator_choice=seperator 
            order.Document_for_prefix=doc_field
            order.Doc_field=f_for_prfx 
            order.seqnum=seq
            order.order_combination=od
            order.save()
        
        messages.success(request, 'Order Generated sequence Updated successfully.')
        return redirect('/number-series-generate/')
    intaial_data = {
        'Field_selection': order.Field_selection
    }
    orderGens = OrderGenerator.objects.all()
    form = orderGeneratorForm(initial=intaial_data,instance=order)
    context = {
        'form':form,
        'orderGens':orderGens,
        'order':order,
    }
    return render(request,'account/Order/update.html', context)  






from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.models import Permission
from django.core import serializers
def roleList(request):
    roles = Group.objects.all()
    context = {
        'roles':roles,
       
    }
    return render(request,'account/Role/role_list.html', context)


def roleCreate(request):
    if request.method == "POST":
        rname = request.POST.get('name')
        perm_list = request.POST.getlist('permissions')
        permission_list_id=[]
        try:
            new_group = Group.objects.filter(name=rname).exists()
            if not new_group:      
                new_group = Group.objects.create(name=rname)
                for perm in perm_list:
                    p=Permission.objects.get(id=int(perm))
                    permission_list_id.append(p.id)
                new_group.permissions.set(permission_list_id) 
                messages.success(request, 'Role created successfully.')
                return redirect('/role-list/') 
        except Exception as e:
            messages.warning(request,'Already exists!.',e)                   
    allperm = Permission.objects.all()
    context = {
        'allperm':allperm
    }
    return render(request,'account/Role/role_create.html',context)





def roleUpdate(request, pk):
    role = get_object_or_404(Group, id=pk)
    if request.method == "POST":
        permission_ids = request.POST.getlist('permissions') 
        #clear the group's existing permissions
        role.permissions.clear()
        
        #Add the selected permissions to the group
        for permmission_id in permission_ids:
            permission = get_object_or_404(Permission, id=permmission_id)
            role.permissions.add(permission)
        
        #save the change to the group
        role.save()    
        messages.success(request,'Role updated successfully!.')
        return redirect('/role-list/')
    
    group = Group.objects.get(name=role)
    perms = group.permissions.all()
    allperm = Permission.objects.exclude(id__in=perms)
    context = {
        'role':role,
        'perms':perms,
        'allperm':allperm,
    }
    return render(request,'account/Role/role_update.html',context)            

def roleDelete(request,pk):
    role = get_object_or_404(Group, pk=pk)
    role.delete()
    return redirect('/role-list/')            





def permissionList(request):
    return render(request, 'account/Permission/permissionbase.html')


def permissionUpdate(request):
    return render(request, 'account/Permission/permission_update.html')




from django.http import HttpResponse
from import_export import resources
from tablib import Dataset

class SalesOrderResource(resources.ModelResource):
    class Meta:
        model = Sales

class SalesDetailsResource(resources.ModelResource):
    class Meta:
        model = Sales_Detail



def export_sales_data(request):
    sales_order_resource = SalesOrderResource()
    sales_details_resource = SalesDetailsResource()

    dataset = Dataset()
    
    # Export SalesOrder data
    queryset_sales_order = Sales.objects.all()
    sales_order_dataset = sales_order_resource.export(queryset_sales_order)
    dataset.extend(sales_order_dataset)

    # # Export SalesDetails data
    # queryset_sales_details = Sales_Detail.objects.all()
    # sales_details_dataset = sales_details_resource.export(queryset_sales_details)
    # dataset.extend(sales_details_dataset)

    response = HttpResponse(dataset.xls, content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="sales_data.xls"'
    
    return response



def export_sales_data(request):
    # Fetch the data from the Sales and Sales_Detail models
    sales_data = Sales.objects.values('id', 'Customer_Name', 'sales_order_no', 'Bill_Date')
    sales_detail_data = Sales_Detail.objects.values('sale_id', 'entity_id', 'Sales_Item_Name', 'Sales_Quantity')

    # Create DataFrames from the data
    sales_df = pd.DataFrame(sales_data)
    sales_detail_df = pd.DataFrame(sales_detail_data)

    # Merge the DataFrames on the 'id' column to create a single DataFrame
    merged_df = pd.merge(sales_df, sales_detail_df, left_on='id', right_on='sale_id')

    # Rename columns as needed
    merged_df.rename(columns={
        'Customer_Name': 'Customer Name',
        'sales_order_no': 'Sales Order No',
        'Bill_Date': 'Sales Order Date',
        'entity_id': 'TID',
        'Sales_Item_Name': 'Sales Item Name',
        'Sales_Quantity': 'Sales Quantity',
    }, inplace=True)

    # Create a response object for CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_data.csv"'

    # Export the DataFrame to CSV
    merged_df.to_csv(response, index=False)

    return response
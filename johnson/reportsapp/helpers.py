from django.template.loader import get_template
from django.http import HttpResponse
from datetime import date ,datetime

from account import models
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill





def sales_to_xlsx():
    """
    Downloads all movies as Excel file with a single worksheet
    """
    sales_queryset = models.Sales.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Sales.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'SALES'

    # Define the titles for columns
    res = [field.name for field in models.Sales._meta.get_fields()]  
    columns = res

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in sales_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            movie.id,
            movie.Bill_Date,
            movie.Invoice_No,
            movie.Tally_MasterID,
            movie.Total_Inventory_Amount,
            movie.Total_GST,
            movie.SGST_AMOUNT,
            movie.CGST_AMOUNT,
            movie.IGST_AMOUNT,
            movie.Cash_Discount_Amount,
            movie.R_O_Amount,
            movie.Total_Invoice_Amount,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def salesreturn_to_xlsx():
    """
    Downloads all movies as Excel file with a single worksheet
    """
    salesreturn_queryset = models.SalesReturn.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Sales.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'SALES'

    # Define the titles for columns
    res = [field.name for field in models.SalesReturn._meta.get_fields()]  
    columns = res

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in salesreturn_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            movie.id,
            movie.Bill_Date,
            movie.Invoice_No,
            movie.Tally_MasterID,
            movie.Total_Inventory_Amount,
            movie.Total_GST,
            movie.SGST_AMOUNT,
            movie.CGST_AMOUNT,
            movie.IGST_AMOUNT,
            movie.Cash_Discount_Amount,
            movie.R_O_Amount,
            movie.Total_Invoice_Amount,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


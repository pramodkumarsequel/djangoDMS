#from dataclasses import field
from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from .models import Event, FieldMaster, ReportsMaster,RoleAssignment
# from .forms import DemoForm
from django.contrib import messages
from .forms import EventForm
from django.db.models import Count, Q
from .helpers import sales_to_xlsx
from django.core.exceptions import ObjectDoesNotExist
import json
from account.models import (Customer, Zone, Distributor, DBPriceMaster, Item, Hierarchy1,
                            Hierarchy2, Hierarchy3, Hierarchy4, Hierarchy5, Sales, Stock,
                            Sales_Detail, PurchaseDocument, DetailOfPurchase, PurchaseReturnDocument,
                            DetailsOfPurchaseReturn, Delivery_Note, Delivery_Note_Details, Receipt_Note,
                            Receipt_Note_Detail, SalesReturn, RoleMaster, User, PurchaseInvoice, salesInvoice, Stock,warehouseStock,warehouseItemChild

                            )
from datetime import datetime

from datetime import timedelta
from django.http import JsonResponse
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.shortcuts import render
from .models import MenuMaster
from .serializers import MenuMasterSerializers,SubMasterSerializers
from django.db.models import Avg, Count, Min, Sum

from .forms import MenuMasterForm, SubMenuForm,EditMenuForm,SubmenuFormset,checkboxform,checkboxformset,RoleAssignmentFilterForm, SearchFilterForm , SearchFilterUpdateForm,editMenuFormsetFactory,itemstockform,dashboardchartForm

from django.contrib.auth.decorators import login_required
from reportsapp.decorators import allowed_users
from account.decorators import unauthenitcated_user
from django.contrib.auth import get_user_model
User = get_user_model()


from decimal import Decimal, ConversionSyntax
from django.utils import timezone
from datetime import datetime  
from django.utils.dateformat import DateFormat
current_date = timezone.now().date()
cdate=DateFormat(current_date)
cd=cdate.format('Y-m-d')




# Create your views here.
def reports_list(request):
    return render(request, 'reportsapp/report_list.html')


def secondary_sale(request):
    if request.method == "POST":
        form = EventForm(request.POST or None)

        print(form)
    form = EventForm()
    return render(request, 'reportsapp/secondarysalebase.html', {'form': form})


from .forms import DocumentMasterForm

import csv

from datetime import datetime

datetime.today().strftime('%Y-%m-%d')


fields = ('Customers', 'Customer_Name','sales_order_no','Bill_Date', 'Invoice_No','cust_po_no','cust_po_date',
                  'Tally_MasterID', 'Total_Inventory_Amount','gst_type', 'Total_GST', 'CGST_AMOUNT', 'SGST_AMOUNT',
                  'Cash_Discount_Amount',
                   'Total_Invoice_Amount',)

from django.apps import apps
def load_fields(request):
    field_id = request.GET.get('fields')
    
    if field_id == "Opening Balance":
        model_meta = warehouseStock._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.IntegerField) if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Sales":
        model_meta=Sales._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Purchase Document": 
        model_meta=PurchaseDocument._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    
    elif field_id == "salesInvoice": 
        model_meta=salesInvoice._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Purchase Return Document": 
        model_meta=PurchaseReturnDocument._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Sales Return": 
        model_meta=SalesReturn._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Delivery Note": 
        model_meta=Delivery_Note._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model  if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Receipt_Note": 
        model_meta=Receipt_Note._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
 
    elif field_id == "PurchaseInvoice":
        model_meta=PurchaseInvoice._meta
        field_names = [field for field in model_meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    else:
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': ""})

    
            

from django.db import models
def load_date_fields(request):
    fields = request.GET.get('fields')
    if fields == "Opening Balance":
        date_fields = [f for f in warehouseStock._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    elif fields == "Sales":
        date_fields = [f for f in Sales._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "SalesInvoice":
        date_fields = [f for f in salesInvoice._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "Delivery Note":
        date_fields = [f for f in Delivery_Note._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
        
    elif fields == "Sales Return":
        date_fields = [f for f in SalesReturn._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "salesInvoice":
        date_fields = [f for f in salesInvoice._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "Purchase Document":
        date_fields = [f for f in PurchaseDocument._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "Receipt_Note":
        date_fields = [f for f in Receipt_Note._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "PurchaseInvoice":
        date_fields = [f for f in PurchaseInvoice._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "Purchase Return Document":
        date_fields = [f for f in PurchaseReturnDocument._meta.get_fields() if isinstance(f, models.DateField) if not isinstance(f, models.DateTimeField)]
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields': date_fields})
    
    elif fields == "":
        return render(request, 'reportsapp/Report/date_fields.html', {'date_fields':''})
        
    
    


def search_report(request):
    if request.method == "GET":
        form  = DocumentMasterForm(request.GET or None)
    elif request.method == "POST":
        form = DocumentMasterForm(request.POST or None)
        doc_type = request.POST.get('Document_Type')
        fld_type = request.POST.get('fields')
        value = request.POST.get('value')
        from_date = request.POST.get('From_Date')
        to_date = request.POST.get('To_Date')      
        if not doc_type:
            messages.error(request, "Please Select Document Type.")    
        if not from_date:
            messages.error(request, "Please select Start Date")
        if not to_date:
            messages.error(request, "Please select End Date.")           
        elif doc_type == "Sales":    
            res = sales_to_xlsx()
            try:
                d1 = Sales.objects.filter(Q(createdAt__range=[from_date, to_date]) and Q(Distributors__Distributor_Code__contains=value) and Q(Distributors__Distributor_Name__contains=value) and Q(Customers__Customer_Code__contains=value) and Q(Customers__Customer_Name__contains=value) and Q(Invoice_No__contains=value))
                return render(request,"reportsapp/salesdatatable.html", {'form': form,'res':res,'Sales':d1})  
            except ObjectDoesNotExist:
                messages.error(request, "Please Select correct Date")                  
        elif doc_type == "Sales Return":
            try:
                d2 = SalesReturn.objects.filter(created_on__range=[from_date, to_date])
                return render(request, "reportsapp/salesreturndatatable.html", {"form":form, "Sales":d2})
            except SalesReturn.DoesNotExist:
                messages.error(request,"Data does not found.")     
        elif doc_type == "Opening Balance":
            try:
                d3 = Stock.objects.all()
                return render(request, "reportsapp/openingbalancedatatable.html", {"form":form, "Sales":d3})
                
            except Stock.DoesNotExist:
                messages.error(request, "Please enter correct Value.")
                res = [field.name for field in Stock._meta.get_fields()]   
                return render(request,"reportsapp/openingbalancedatatable.html", {'form':form, 'Sales':d3})     
        elif doc_type == "Purchase Document":
            res = [field.name for field in PurchaseDocument._meta.get_fields()]
            try:
                d4=PurchaseDocument.objects.filter(created_on__range=[from_date, to_date])
                return render(request, "reportsapp/purchasedatatable.html",{"form":form, "Sales":d4})
            except PurchaseDocument.DoesNotExist:
                messages.error(request, "Value Does not Exist.")
        elif doc_type == "Purchase Return Document":
            try:     
                res = [field.name for field in PurchaseReturnDocument._meta.get_fields()]   
                d5 = PurchaseReturnDocument.objects.filter(created_on__range=[from_date, to_date])
                return render(request, "reportsapp/purchasereturndatatable.html", {'from':form, 'Sales':d5})
            except PurchaseReturnDocument.DoesNotExist:
                messages.error(request,"Value Does not Exists.")
        elif doc_type == "Delivery Note":
            res = [field.name for field in Delivery_Note._meta.get_fields()]
            try:
                D6 = Delivery_Note.objects.filter(created_on__range=[from_date, to_date])
                return render(request, "reportsapp/deliverydatatable.html",{'form':form,"Sales":D6})
            except Delivery_Note.DoesNotExist:
                messages.error("Value Does not exists.")
        elif doc_type == "Receipt Note":
            res = [field.name for field in Receipt_Note._meta.get_fields()]
            try:  
                D7 = Receipt_Note.objects.filter(created_on__range=[from_date, to_date])
                return render(request, "reportsapp/receiptdatatable.html",{"form":form, "Sales":D7})
            except Receipt_Note.DoesNotExist:
                messages.error(request, "Value Does not exist.")
                 
    form = DocumentMasterForm()    
    return render(request, 'reportsapp/searchreportbase.html', {'form': form})



def export_sales_to_xlsx(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    sales_queryset = Sales.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Sales.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'SALES'

    # Define the titles for columns
    columns = [
        'ID',
        'Distributor Code',
        'Distributor Name',
        'Customer Code',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in sales_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            movie.id,
            movie.CGST_AMOUNT,
            movie.IGST_AMOUNT,
            movie.Invoice_No,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def my_view(request, queryset, fields):
    response = HttpResponse(mimetype='text/csv')
    response['Content-Disposition'] = 'attachment;filename=export.csv'
    writer = csv.writer(response)
    for obj in queryset:
        writer.writerow([getattr(obj, f) for f in fields])
    return response


from django.contrib.auth.models import User
from django.shortcuts import render
from .filters import CustomerFilter


def search(request):
    customer_list = Customer.objects.all()
    customer_filter = CustomerFilter(request.GET, queryset=customer_list)
    return render(request, 'search/user_list.html', {'filter': customer_filter})


from django.db.models import Sum, Count
import matplotlib

matplotlib.use('Agg')
from matplotlib import pyplot as plt
import numpy as np

import random

from collections import OrderedDict 
def dashboard(request):
    items = Item.objects.values('Item_Code').distinct().count()
    customers = Customer.objects.values('Customer_Name').distinct().count()
    qs = Sales_Detail.objects.aggregate(Sum('Total_Amount'))
    items_MRP = Item.objects.aggregate(Sum('MRP'))
    # ob = DistributorModel.objects.values('DistributorName').distinct().count()
    # partstate = DistributorModel.objects.values('State').annotate(Sum('Amount'))[:3]
    # res = DistributorModel.objects.values('DistributorName').annotate(Count('DocID')).order_by().filter(DocID__gt=1)
    # res1 = DistributorModel.objects.filter(DistributorName__in=[item['DistributorName'] for item in res])
    #
    # queryset = DistributorModel.objects.order_by('-DistributorName')[:5]
    sales_data = Sales.objects.values('Total_Invoice_Amount').annotate(Count('Total_Inventory_Amount'))
    dataset = Item.objects.values('Item_Name', 'Item_Code').annotate(Item_MRP=Sum('MRP')).order_by('Item_Name')
    dataset_pd = pd.DataFrame(dataset)
    # distributor_data = Distributor.objects.values('Distributor_Name', 'Distributor_Code')
    distributor_data = Sales_Detail.objects.values('Sales_Item_Name').annotate(Count('id'))
    return_dat = [list(q) for q in distributor_data]
    salesVal = qs['Total_Amount__sum']
    itm_mrp = items_MRP['MRP__sum']
    labels = 'Sale', 'Items'

    return render(request, 'reportsapp/dashboard.html',
                  {'customers': customers, 'items': items, 'salesval': salesVal, 'itm_mrp': itm_mrp, 'dataset':dataset,
                   'dataset4': "distributor_data",
                   })




# Create your views here.
 
def dynamicmenu(request):
    try:            
        menuList = MenuMaster.objects.all().filter(Pmenu=0)
        mainmenu = MenuMasterSerializers(menuList,many=True)
        submenulist = MenuMaster.objects.all().filter(Pmenu__in=[1,2,3,4,5,6,7,8,9,10,11,12])
        data = mainmenu.data
        request.session['mainM'] = data
        print(request.session['mainM'])
 
        submenudata = SubMasterSerializers(submenulist,many=True)
        subdata = submenudata.data    
        request.session['submenu'] = subdata
        print(request.session['submenu'])
        return render(request, 'reportsapp/index1.html', {})  
 
    except Exception as identifier:         
        return render(request, 'reportsapp/index1.html', {}) 
    

@login_required(login_url='login')

def menumaster(request):
    currentUrl = request.get_full_path()
    formset = SubmenuFormset()
    editmenuform = EditMenuForm()
    if request.method == "POST":
        isview = request.POST.get('Is_View')
        iscreate = request.POST.get('Is_Create')
        isedit = request.POST.get('is_Edit')
        isdelete = request.POST.get('is_Delete')
        hidden = request.POST.get('hidden')
        usrole = request.POST.get('usrole')
        print(usrole)
        
        try:
            rolename=RoleMaster.objects.get(id=isview)
        except RoleMaster.DoesNotExist:
            rolename=None
    if request.method == "POST" and 'rootmenu' in request.POST:
        form = MenuMasterForm(request.POST or None)
        try:
            if form.is_valid():
                dm=form.save()
                menu_name = form.cleaned_data['MenuName']
                dm.pagelink = f'{currentUrl}{"="}{menu_name}'

                # dm.save()
                messages.success(request, "Root Menu Add successfully.")
            else:
                form = MenuMasterForm()
        except Exception as e:
            return e
    if request.method == "POST"  and 'submenu' in request.POST:
        formset = SubMenuForm(request.POST)
        try:
            if formset.is_valid():  
                pmenu = formset.cleaned_data['Pmenu']
                pglinktype = formset.cleaned_data['pagelinktype']
                pglink = formset.cleaned_data['pagelink']
                menuname = formset.cleaned_data['MenuName']
                dord = formset.cleaned_data['Dord']
                # MenuMaster(
                #     Pmenu=pglink,
                #     pagelink = pglink,
                #     MenuName=menuname,
                #     Dord = dord            
                # ).save()
                messages.success(request, "Sub Menu Add successfully.")
                return redirect('/report/menumaster/')
                
                 
            else:
                formset = SubMenuForm()
        except Exception as e:  
            return e      
        
    form=MenuMasterForm()
    formset = SubMenuForm()
    checkform = checkboxformset()
    roles = RoleMaster.objects.all().order_by('name')
    menus = MenuMaster.objects.all()
    context = {
        'roles':roles,
        'form':form,
        'form1':formset,
        'menus':menus,
        'editmenuform': editmenuform,
        'checkform':checkform,
    }
    return render(request, 'reportsapp/menumaster/menumaster.html',context)    


def checkbox(request):
    if request.method == "POST":
        isview = request.POST.get('Is_View')
        iscreate = request.POST.get('Is_Create')
        isedit = request.POST.get('is_Edit')
        isdelete = request.POST.get('is_Delete')
    return render(request,'reportsapp/menumaster/menumaster.html')

def load_user(request):
    user_id = request.GET.get('username')    
    users = User.objects.filter(id=user_id).order_by('username')
    context = {'users': users}
    return render(request, 'account/user_dropdown.html', context)    


from account.models import User
from .forms import DistributorMultiChoiceForm

from .decorators import menu_show_as_per_role,check_perm_required
@login_required(login_url='login')

def role_assignment_create(request):
    if request.method == "POST":
        roleid = request.POST.get('ROLE')
        userid = request.POST.get('User')
        docmnttype = request.POST.get('doc_type')
        form = SearchFilterForm(request.POST)
        if form.is_valid():
            dis = form.cleaned_data.get("distributors")
            zones = form.cleaned_data.get('zones')
            doctype = form.cleaned_data.get('documents')
            Can_Create = form.cleaned_data.get('iscreate')
            Can_Edit = form.cleaned_data.get('isedit')
            Can_View = form.cleaned_data.get('isview')
            current_user_id = request.user.id
            roleasign=RoleAssignment.objects.create(
                EID=current_user_id,
                UID=userid,
                rolename = roleid,
                documenttype = docmnttype,
                iscreate=Can_Create,
                isedit=Can_Edit,
                isview=Can_View,
            )
            roleasign.save()
            roleasign.distributors.set(dis)
            roleasign.zones.set(zones)
            roleasign.documents.set(doctype)
            roleasign.save()
            messages.success(request, "record created successfully!")
    doctype = ReportsMaster.objects.filter()        
    roles = RoleMaster.objects.exclude(name="SU")
    distributors = Distributor.objects.all()
    zones = Zone.objects.all()
    users = User.objects.all()
    form = RoleAssignmentFilterForm()
    dis = DistributorMultiChoiceForm()
    ser = SearchFilterForm()
    context = {
        'doctype':doctype,
        'roles':roles,
        'users':users,
        'distributors':distributors,
        'zones':zones,
        'form':form,
        'distributors1':dis,
        'ser':ser,

    }
    return render(request, 'reportsapp/role_assignment.html',context)

from .decorators import role_required
def roleassignment_update(request, pk):
    data = get_object_or_404(RoleAssignment,id=pk)
    if request.method == "POST":
        serform = SearchFilterUpdateForm(request.POST,instance=data)
        if serform.is_valid():    
            serform.save()
            return redirect('/role_assignment/')

    serform = SearchFilterUpdateForm(instance=data)          
    doctype = ReportsMaster.objects.all()       
    roles = RoleMaster.objects.exclude(name="SU")
    distributors = Distributor.objects.all()
    zones = Zone.objects.all()
    users = User.objects.all()
    form = RoleAssignmentFilterForm()
    dis = DistributorMultiChoiceForm()
    context = {
        'doctype':doctype,
        'roles':roles,
        'users':users,
        'distributors':distributors,
        'zones':zones,
        'serform':serform,
        'distributors1':dis,
        'data': data,

    }    
    return render(request, 'reportsapp/roleassignment_update.html',context)



def menumaster1(request):
    if request.method == "POST":
        mtype = request.POST.get('mt')
        Mname = request.POST.get('M')
        diO = request.POST.get('Dord', None)
        isv = request.POST.getlist('IsView')
        isc = request.POST.getlist('IsCreate')
        ise = request.POST.getlist('isEdit')
        isd = request.POST.getlist('isDelete')
        h = request.POST.getlist('content')
        rol = request.POST.getlist('role')
        res=MenuMaster.objects.create(
                        MenuName=Mname,
                        MTYPE = mtype,
                        Dord = diO,
                        Roles = [isv, isc, isd,ise]
                                             
        )
        res.save()
        
        
            
            
 
    form = MenuMasterForm()
    formset = editMenuFormsetFactory()
    context = {
        'form':form,
        'roles':RoleMaster.objects.all(),
        'formset':formset
    }
    return render(request,'reportsapp/menumaster/menumaster1.html',context)

def RoleAssignUser(request):
    roleid = request.GET.get('roleid','')
    rolename = RoleMaster.objects.get(id=roleid)
    users = User.objects.filter(role_id=roleid)
    docs = ReportsMaster.objects.none
    return render(request, 'reportsapp/menumaster/user_with_role.html', {'users': users})
    #return HttpResponse(json.dumps({'uom': base_uom}), content_type="application/json") 

def AssignUser(request):
    user_id = request.GET.get('userrole')
    role_id = request.GET.get('roleid')
    rolename = RoleMaster.objects.get(id=role_id)

    try:
        if str(rolename) == str(user_id): 
            docs = ReportsMaster.objects.filter(id=1)
            res = ",".join([str(i) for i in docs])
            return render(request, 'reportsapp/menumaster/doctype_dropdown.html', {'docs':res})
        else:
            docs = ReportsMaster.objects.none()        
        return render(request, 'reportsapp/menumaster/doctype_dropdown.html')
    except User.DoesNotExist:
        user = None   
    docs = ReportsMaster.objects.none
    return render(request, 'reportsapp/menumaster/doctype_dropdown.html') 


def RoleAssignmentFilter(request):
    form =  RoleAssignmentFilterForm()
    return render(request,'reportsapp/menumaster/RoleAssignment_filter.html',{'form': form})      
    


def base1(request):
    items = Item.objects.values('Item_Code').distinct().count()
    customers = Customer.objects.values('Customer_Name').distinct().count()
    qs = Sales_Detail.objects.aggregate(Sum('Total_Amount'))
    items_MRP = Item.objects.aggregate(Sum('MRP'))
    # ob = DistributorModel.objects.values('DistributorName').distinct().count()
    # partstate = DistributorModel.objects.values('State').annotate(Sum('Amount'))[:3]
    # res = DistributorModel.objects.values('DistributorName').annotate(Count('DocID')).order_by().filter(DocID__gt=1)
    # res1 = DistributorModel.objects.filter(DistributorName__in=[item['DistributorName'] for item in res])
    #
    # queryset = DistributorModel.objects.order_by('-DistributorName')[:5]
    sales_data = Sales.objects.values('Total_Invoice_Amount').annotate(Count('Total_Inventory_Amount'))
    dataset = Item.objects.values('Item_Name').annotate(Item_MRP=Sum('MRP')).order_by('Item_Name')
    # distributor_data = Distributor.objects.values('Distributor_Name', 'Distributor_Code')
    distributor_data = Sales_Detail.objects.values('Sales_Item_Name').annotate(Count('id'))
    return_dat = [list(q) for q in distributor_data]
    salesVal = qs['Total_Amount__sum']
    itm_mrp = items_MRP['MRP__sum']
    purchases = PurchaseDocument.objects.all().count()
    sales = Sales.objects.all().count()
    labels = 'Sale', 'Items'
    
    context = {'customers': customers, 'items': items, 'salesval': salesVal, 'itm_mrp': itm_mrp, 'dataset': dataset, "purchases":purchases, "sales":sales,} 
    return render(request, 'reportsapp/base1.html',context)        



from decimal import Decimal, InvalidOperation
from account.models import Sales_Return_Detail,saleinvoicelineItem,PurchaseinvoicelineItem,warehouseItemChild,DetailsOfPurchaseReturn,Delivery_Note_Details
# search reports

def get_date_format(object):
    if object == None:
        return cd
    dm=DateFormat(object)
    dm_format = dm.format('Y-m-d')
    return dm_format



error_message = "ValueError: string must be a decimal or a number convertible to Decimal."
from itertools import chain
def searchReport(request):
    if request.method == "GET":
        form = DocumentMasterForm(request.GET or None)
    elif request.method == "POST":  
        form = DocumentMasterForm(request.POST or None)
        doc_type = request.POST.get('Document_Type')
        fld_type = request.POST.get('fields')
        fld_value = request.POST.get('value')
        date_type = request.POST.get('Date_Type')
        from_date = request.POST.get('From_Date')
        to_date = request.POST.get('To_Date') 
        if doc_type:
            fields = [field.verbose_name for field in Sales._meta.get_fields() if not isinstance(field, models.BooleanField) if not field.related_model if not isinstance(field, models.DateTimeField)]+[field.verbose_name for field in Sales_Detail._meta.get_fields() if not field.related_model if not isinstance(field, models.BooleanField)]
            if doc_type=="Sales":
                try:
                    if fld_value and fld_value:
                        filter_query = {}
                        filter_query[fld_type] = fld_value
                        sal_qery = Sales.objects.filter(**filter_query)
                        child_queryset = [d for f in list(sal_qery) for d in Sales_Detail.objects.filter(entity_id=f.id)]
                        return render(request, 'reportsapp/Report/partial_report_datatable.html', {'queryset':sal_qery,'fields':fields,'form':form,'queryset2':child_queryset, 'field_names':fields}) 
                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select field name and value.',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                        

                    query_res = Sales.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    child_query = [f for f in list(query_res) for f in Sales_Detail.objects.filter(entity_id=f.id)]
                    return render(request, 'reportsapp/Report/partial_report_datatable.html',{'form':form,'queryset':query_res,'field_names':fields, 'queryset2':child_query})
                        
                except Exception as e:
                    messages.error(request, f'{e}', extra_tags='alert-danger')
                    return redirect('/report/searchReport/')
                          
            elif doc_type=="Purchase Document":
                field_names = [field for field in PurchaseDocument._meta.get_fields() if not isinstance(field, models.BooleanField) if not field.related_model if not isinstance(field, models.DateTimeField)]+[field for field in DetailOfPurchase._meta.get_fields() if not field.related_model if not isinstance(field, models.BooleanField)]   
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        qury = PurchaseDocument.objects.filter(**filter_condition)
                        queryset_child = [obj for data in list(qury) for obj in DetailOfPurchase.objects.filter(entity_id=data.id)]
                        return render(request, 'reportsapp/Report/partial_purchase_report_datatable.html',{'form':form, 'queryset':qury, 'queryset2':queryset_child, 'field_names':field_names})

                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select field type and value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                        
                    query_res = PurchaseDocument.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    child_query = [f for f in list(query_res) for f in DetailOfPurchase.objects.filter(entity_id=f.id)] 
                    return render(request, 'reportsapp/Report/partial_purchase_report_datatable.html',{'form':form,'queryset':query_res,  'field_names':field_names})
   
                except Exception as e:
                    messages.error(request, f'{e}', extra_tags="alert-danger")
                    return redirect('/report/searchReport/')
                    
      
            elif doc_type=="Receipt_Note":
                fields = [field for field in Receipt_Note._meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.BooleanField)] + [f for f in Receipt_Note_Detail._meta.get_fields() if not f.related_model if not f.name == "MRN_QTY"]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = Receipt_Note.objects.filter(**filter_condition)
                        return render(request, 'reportsapp/Report/partial_receipt_note_datatable.html',{'fields':fields,'form':form, 'queryset':queryset})
                    
                    elif fld_type or fld_value:  
                        messages.warning(request, 'Please select Field Type and Field Value.',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    queryset = Receipt_Note.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    return render(request, 'reportsapp/Report/partial_receipt_note_datatable.html',{'fields':fields,'form':form, 'queryset':queryset})
         
                except Exception as e:
                    messages.error(request, f'{e}', extra_tags="alert-danger")
                    return redirect('/report/searchReport/')  


            elif doc_type=="PurchaseInvoice":
                fields = [field for field in PurchaseInvoice._meta.get_fields() if not field.name == "isactive" if not field.related_model if not isinstance(field, models.DateTimeField)] + [f for f in PurchaseinvoicelineItem._meta.get_fields() if not f.related_model if not f.name=="PI_QTY"]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = PurchaseInvoice.objects.filter(**filter_condition)
                        return render(request, 'reportsapp/Report/partial_purchaseinv_datatable.html',{'fields':fields,'form':form,'queryset':queryset})
                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')   
                    
                    queryset = PurchaseInvoice.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date) 
                    return render(request, 'reportsapp/Report/partial_purchaseinv_datatable.html',{'fields':fields,'form':form,'queryset':queryset})      

                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/searchReport/')


            elif doc_type == "Opening Balance":
                fields = [field for field in warehouseStock._meta.get_fields() if not field.related_model if not field.name =="id" if not field.name == "itemName" if not field.name =="isactive" if not field.name == "Date" if not isinstance(field, models.DateTimeField)  if not isinstance(field, models.BooleanField)]+[field for field in warehouseItemChild._meta.get_fields() if not field.related_model if not field.name =="Batch" if not field.name =="Reference_No" if not field.name =="w_house"]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = warehouseStock.objects.all()
                        return render(request, 'reportsapp/Report/partial_warehousestock_datatable.html',{'fields':fields,'form':form,'queryset':queryset})
                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    queryset = warehouseStock.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date) 
                    return render(request, 'reportsapp/Report/partial_warehousestock_datatable.html',{'fields':fields,'form':form,'queryset':queryset})

                    
                        
                except Exception as e:
                    print(e)
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/searchReport/')                         
                                            
                    
            elif doc_type == "Purchase Return Document":
                fields=[f for f in PurchaseReturnDocument._meta.get_fields() if not f.name =="isactive" if not f.related_model if not isinstance(f, models.DateTimeField)]+[f for f in DetailsOfPurchaseReturn._meta.get_fields() if not f.name =="Reference_No" if not f.related_model]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset =  PurchaseReturnDocument.objects.filter(**filter_condition)  
                        return render(request, 'reportsapp/Report/partial_purchasereturn_datatable.html',{'fields':fields, 'form':form, 'queryset':queryset})

                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    queryset = PurchaseReturnDocument.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    return render(request, 'reportsapp/Report/partial_purchasereturn_datatable.html',{'fields':fields, 'form':form, 'queryset':queryset})


                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-error")
                    return redirect('/report/searchReport/')  


            elif doc_type == "Delivery Note":
                fields=[f for f in Delivery_Note._meta.get_fields() if not f.related_model if not isinstance(f, models.BooleanField) if not isinstance(f, models.DateTimeField) if not f.name=="Bill_Date" if not f.name=="Invoice_No"] + [f for f in Delivery_Note_Details._meta.get_fields() if not f.related_model if not f.name=="DN_QTY" if not f.name == "isactive"]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = Delivery_Note.objects.filter(**filter_condition)
                        return render(request, 'reportsapp/Report/partial_deliverynote_datatable.html',{'fields':fields, 'form':form,'queryset':queryset})

                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value.',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    
                    queryset = Delivery_Note.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    return render(request, 'reportsapp/Report/partial_deliverynote_datatable.html',{'fields':fields, 'form':form,'queryset':queryset})
          
                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/searchReport/') 
                    
            elif doc_type == "Sales Return":
                fields=[f for f in SalesReturn._meta.get_fields() if not f.name == "isactive" if not f.name == "Tally_Master_ID" if not f.related_model if not isinstance(f, models.DateTimeField)] + [f for f in Sales_Return_Detail._meta.get_fields() if not f.related_model]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = SalesReturn.objects.filter(**filter_condition)
                        return render(request, 'reportsapp/Report/partial_salereturn_datatable.html',{'fields':fields, 'form':form, 'queryset':queryset})


                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    
                    queryset = SalesReturn.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    return render(request, 'reportsapp/Report/partial_salereturn_datatable.html',{'fields':fields, 'form':form, 'queryset':queryset})

                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/searchReport/') 
 
            elif doc_type == "salesInvoice":
                fields=[f for f in salesInvoice._meta.get_fields() if not f.related_model if not f.name=="isactive" if not isinstance(f, models.DateTimeField)] + [f for f in saleinvoicelineItem._meta.get_fields() if not f.related_model if not f.name =="SI_QTY"]
                try:
                    if fld_type and fld_value:
                        filter_condition = {}
                        filter_condition[fld_type] = fld_value
                        queryset = salesInvoice.objects.filter(**filter_condition)
                        return render(request, 'reportsapp/Report/partial_salesinvoice_datatable.html', {'fields':fields, 'form':form,'queryset':queryset})
                    
                    elif fld_type or fld_value:
                        messages.warning(request, 'Please select Field Type and Field Value',extra_tags="alert-warning")
                        return redirect('/report/searchReport/')
                    
                    queryset = salesInvoice.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date)
                    return render(request, 'reportsapp/Report/partial_salesinvoice_datatable.html', {'fields':fields, 'form':form,'queryset':queryset})
             
                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/searchReport/') 
        
        messages.error(request, 'Please select Document Type',extra_tags="alert-danger")
        return redirect('/report/searchReport/')            
    form = DocumentMasterForm()        
    context = {
        'form':form
    }
    return render(request, 'reportsapp/Report/reportbase.html', context)



def transaction(request):
    return render(request, 'reportsapp/Transaction/transactionbase.html')



def clientData(request):
    sale_data_points=[]
    purchase_data_points=[]
    w_data_points=[]
    sales_price=[]
    customers=Customer.objects.all().count()
    items=Item.objects.all().count()
    sales = Sales.objects.all()
    purchases=PurchaseDocument.objects.all()
    deliverys = Delivery_Note.objects.all()
    ws=warehouseItemChild.objects.all()
    
    for sale in sales:
        sale_data_points.append({
            'label': sale.Customers.Customer_Name,
            'y':int(sale.Total_Invoice_Amount)
        })
    
    for pur in purchases:
        purchase_data_points.append({
            'label':pur.Distributors.Distributor_Name,
            'y':int(pur.Total_Invoice_Amount)
        }) 
        
    for w in ws:
        w_data_points.append({
            'label':w.Item_Code.Item_Name,
            'y':int(w.Quantity)
        }) 
              
    context = {
        'customers':customers,
        'items':items,
        'sale_data_points':sale_data_points,
        'purchase_data_points':purchase_data_points,
        'w_data_points':w_data_points,
        
    }
    return render(request, 'reportsapp/Transaction/sales_client.html', context)    





def saleDashboard(request):
    return render(request, 'reportsapp/Transaction/sale_dashboard.html')


from account.forms import warehouseclosingBalance, Warehouse
from account.global_dopdown import load_itemname_stock,load_stock_wiseitem,load_warehousewise_item
from django.http import HttpResponseServerError
from django.db.models import DecimalField, IntegerField, ExpressionWrapper, Max , Sum , F
from django.db.models.functions import Coalesce

def closingbalanceReport(request):
    query=None
    try:
        if request.method == "GET":
            form = warehouseclosingBalance(request.GET or None)
        elif request.method == "POST":
            warehouse_code=request.POST.get('w_name') 
            warehouse = Warehouse.objects.get(id=warehouse_code)
            ware_id = warehouseStock.objects.filter(warehouseCode_id=warehouse.id)
            items=request.POST.get('items')
            date=request.POST.get('date')
            if warehouse_code and items:
                q =  warehouseItemChild.objects.filter(Item_Code_id=items).values('id', 'Quantity')
                for d in ware_id:
                    query =  warehouseItemChild.objects.filter(Item_Code_id=items).filter(warehouseID=d.id).values(
                                                                                        'id',
                                                                                        'Quantity',
                                                                                        'Item_Code__Item_Name',
                                                                                        'Item_Code__sales_detail',
                                                                                        'Item_Code__sales_detail__Sales_Quantity',
                                                                                        'Item_Code__sales_return_detail',
                                                                                        'Item_Code__delivery_note_details__Quantity',
                                                                                        'Item_Code__sales_return_detail__Sales_Quantity',
                                                                                        'Item_Code__detailofpurchase__Purchases_Quantity',
                                                                                        'Item_Code__receipt_note_detail__Quantity',
                                                                                        'Item_Code__detailsofpurchasereturn__Purchases_Return_Quantity',
                                                                                        'Item_Code__Item_Code',
                                                                                        'Item_Code__Hierarchy1__Name',
                                                                                        'Item_Code__Hierarchy2__Name',
                                                                                        'Item_Code__Hierarchy3__Name',
                                                                                        'Item_Code__Hier4__Name',
                                                                                        'Item_Code__Hier5__Name'
                                                                                        #).annotate(closing_balance=F('Quantity') + Coalesce(Sum(F('Item_Code__detailofpurchase__Purchases_Quantity')),0)+ Coalesce(Sum(F('Item_Code__sales_return_detail__Sales_Quantity')),0) - Coalesce(Sum(F('Item_Code__detailsofpurchasereturn__Purchases_Return_Quantity')),0) -(Coalesce(Sum(F('Item_Code__sales_detail__Sales_Quantity')),0)))
                                ).annotate(closing_balance=F('Quantity') + Coalesce(Sum(F('Item_Code__detailofpurchase__Purchases_Quantity')),0) +Coalesce(Sum(F('Item_Code__sales_return_detail__Sales_Quantity')),0) -Coalesce(Sum(F('Item_Code__delivery_note_details__Quantity')),0) - Coalesce(Sum(F('Item_Code__receipt_note_detail__Quantity')),0) -(Coalesce(Sum(F('Item_Code__sales_detail__Sales_Quantity')),0)))

                form=warehouseclosingBalance()
                return render(request, 'reportsapp/closingstock/partial_closingbalance_datatable.html',{'form':form,'sales':query, 'warehouses':ware_id})  
                
            if date > cd:
                messages.warning(request, 'Data not found.')
                return redirect('/report/closing_balance/')
            elif date <= cd: 
                qm =  warehouseItemChild.objects.filter(created_on__lte=date).values(
                                                                                    'id',
                                                                                    'Quantity',
                                                                                    'Item_Code__TaxRate',
                                                                                    'Item_Code__Item_Name',
                                                                                    'Item_Code__sales_detail',
                                                                                    'Item_Code__sales_detail__Sales_Quantity',
                                                                                    #'Item_Code__delivery_note_details__Quantity',
                                                                                    'Item_Code__sales_return_detail',
                                                                                    'Item_Code__sales_return_detail__Sales_Quantity',
                                                                                    'Item_Code__detailofpurchase__Purchases_Quantity',
                                                                                    'Item_Code__detailsofpurchasereturn__Purchases_Return_Quantity',
                                                                                    'Item_Code__Item_Code',
                                                                                    'Item_Code__Hierarchy1__Name',
                                                                                    'Item_Code__Hierarchy2__Name',
                                                                                    'Item_Code__Hierarchy3__Name',
                                                                                    'Item_Code__Hier4__Name',
                                                                                    'Item_Code__Hier5__Name'
                                                                                    ).annotate(closing_balance=F('Quantity') + Coalesce(Sum(F('Item_Code__detailofpurchase__Purchases_Quantity')),0) -(Coalesce(Sum(F('Item_Code__sales_detail__Sales_Quantity')),0) - Coalesce(Sum(F('Item_Code__sales_return_detail__Sales_Quantity')),0)))
                form=warehouseclosingBalance()
                return render(request, 'reportsapp/closingstock/partial_closingbalance_datatable.html',{'form':form,'sales':qm})
   
    except Exception as e:
        return HttpResponseServerError("An error occurred: " + str(e))    
        # messages.warning(request, 'data not found.') 
        # return redirect('/report/closing_balance/')

    form = warehouseclosingBalance() 
    context = {
        'form':form,
        'items':Item.objects.all(),
    }   
    return render(request, 'reportsapp/closingstock/partial_closingbalance_datatable.html',context)

from reportsapp.forms import itemstockform
from reportsapp.models import stockstatement
from itertools import chain
def itemwiseStock(request):
    if request.method == "POST":
        form = itemstockform(request.POST or None)
        itmcode = request.POST.get('itemcode')
        itmname = request.POST.get('itemname')
        date = request.POST.get('date')
        print(itmname)
        if date <= cd and itmcode:
            try:
                queryset=warehouseItemChild.objects.filter(created_on__lte=date).values(
                                                                                    'id',
                                                                                    'Quantity',
                                                                                    'Item_Code__TaxRate',
                                                                                    'Item_Code__Item_Name',
                                                                                    'Item_Code__sales_detail',
                                                                                    'Item_Code__sales_detail__Sales_Quantity',
                                                                                    'Item_Code__delivery_note_details__Quantity',
                                                                                    'Item_Code__sales_return_detail',
                                                                                    'Item_Code__sales_return_detail__Sales_Quantity',
                                                                                    'Item_Code__detailofpurchase__Purchases_Quantity',
                                                                                    'Item_Code__detailsofpurchasereturn__Purchases_Return_Quantity',
                                                                                    'Item_Code__Item_Code',
                                                                                    'Item_Code__Hierarchy1__Name',
                                                                                    'Item_Code__Hierarchy2__Name',
                                                                                    'Item_Code__Hierarchy3__Name',
                                                                                    'Item_Code__Hier4__Name',
                                                                                    'Item_Code__Hier5__Name'
                                                                                    ).annotate(closing_balance=F('Quantity') + Coalesce(Sum(F('Item_Code__detailofpurchase__Purchases_Quantity')),0) -(Coalesce(Sum(F('Item_Code__sales_detail__Sales_Quantity')),0) - Coalesce(Sum(F('Item_Code__sales_return_detail__Sales_Quantity')),0)))
                context = {
                    'form':form,
                    'sales':queryset,
                    
                }
                return render(request, 'reportsapp/itemstock/partial_item_stock_datatable.html',context)

                
            except Exception as e:
                messages.warning(request,'Already exists!.',e)     
        
        messages.warning(request, 'Data not found.')
        return redirect('/report/itemwiseStock/')        
    
    else:
        form=itemstockform()
    context = {
            'form':form,
        }
    return render(request, 'reportsapp/itemstock/partial_item_stock_datatable.html', context)


from reportsapp.utils import get_chart  
from datetime import datetime
from django.forms import model_to_dict
  
import pandas as pd


def load_document_fields(request):
    field_id = request.GET.get('fields')
    
    if field_id == "Opening Balance":
        model_meta = warehouseStock._meta
        field_names = [field for field in model_meta.get_fields() if not field.name =="isactive" if not field.related_model if not isinstance(field, models.IntegerField) if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in warehouseItemChild._meta.get_fields() if not field.related_model if not field.name=="warehouseID" if not field.name=="id" if not isinstance(field, models.DateTimeField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Sales":
        model_meta=Sales._meta
        field_names = [field for field in model_meta.get_fields() if not field.name =="isactive" if not field.name =="id" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in Sales_Detail._meta.get_fields() if not field.name == "id"  if not field.name=="isactive" if not field.related_model if not field.name =="entity_id" if not isinstance(field, models.DateTimeField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "PurchaseDocument": 
        model_meta=PurchaseDocument._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in DetailOfPurchase._meta.get_fields() if not isinstance(field, models.DateTimeField) if not field.related_model if not field.name=="id" if not field.name=="entity_id"]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    
    elif field_id == "salesInvoice": 
        model_meta=salesInvoice._meta
        field_names = [field for field in model_meta.get_fields() if not field.name =="isactive" if not field.related_model if not field.name =="id" if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in saleinvoicelineItem._meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not field.name=="id" if not field.name=="parent_id" if not field.name=="SI_QTY"]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "PurchaseReturnDocument": 
        model_meta=PurchaseReturnDocument._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in DetailsOfPurchaseReturn._meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not field.name=="id" if not field.name=="entity_id"]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "SalesReturn": 
        model_meta=SalesReturn._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in Sales_Return_Detail._meta.get_fields() if not field.related_model if not field.name=="id" if not field.name=="entity_id" if not isinstance(field, models.DateTimeField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Delivery_Note": 
        model_meta=Delivery_Note._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model  if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in Delivery_Note_Details._meta.get_fields() if not field.related_model if not isinstance(field, models.DateTimeField) if not field.name=="id" if not field.name=="entity_id" if not field.name=="DN_QTY" if not field.name=="isactive"]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
    elif field_id == "Receipt_Note": 
        model_meta=Receipt_Note._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in Receipt_Note_Detail._meta.get_fields() if not field.related_model if not field.name=="id" if not field.name=="entity_id" if not field.name=="MRN_QTY" if not isinstance(field, models.DateTimeField)]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    
 
    elif field_id == "PurchaseInvoice":
        model_meta=PurchaseInvoice._meta
        field_names = [field for field in model_meta.get_fields() if not field.name=="id" if not field.name =="isactive" if not field.related_model if not isinstance(field, models.DateTimeField) if not isinstance(field, models.DateField)]+[field for field in PurchaseinvoicelineItem._meta.get_fields() if not field.related_model if not field.name =="id" if not field.name=="parent_id" if not isinstance(field, models.DateTimeField) if not field.name=="PI_QTY"]
        return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': field_names})
    

    return render(request, 'reportsapp/fields_dropdown_list_options.html', {'fields': ""})


def filter_fields(model, field_names):
    model_fields = model._meta.get_fields()
    filtered_fields = []

    for field_name in field_names:
        for field in model_fields:
            if field.name == field_name:
                filtered_fields.append(field)
                break
    
    return filtered_fields 
       

         
def dashboardChart(request):
    sales_df =None
    merged_df =None
    positions_df = None
    df =None
    chart = None
    no_data = None
    form=dashboardchartForm(request.POST or None)
    if request.method == "POST":
        doc_type=request.POST.get('document_type')
        chart_type=request.POST.get('chart_type')
        from_date=request.POST.get('from_date')
        to_date=request.POST.get('to_date')
        multiselect = request.POST.getlist('groups[]')   
        if doc_type:
            if doc_type == "Sales":
                data_points = []
                data_point = []
                try:
                    all_fields =[f.verbose_name for f in Sales._meta.get_fields() if not f.related_model if not isinstance(f, models.FloatField) if not isinstance(f, models.BooleanField)]
                    qs = Sales.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('sales_detail_set') 
                    if len(qs)>0: 
                        for data in qs:
                            obj = data.sales_detail_set.all()
                            total_sum = obj.aggregate(total_sales=Sum('Total_Amount'))['total_sales']
                            data_points.append({
                                'label':data.Customers.Customer_Name,
                                'y':str(total_sum)
                            })
                        form = dashboardchartForm()    
                        return render(request, 'reportsapp/dashboard/partial_sale_datatable.html',{'sales':qs, 'form':form, 'fields':all_fields, 'chart':chart_type,'data_points':data_points, 'Doc_type':doc_type})            
                    else:
                        messages.warning(request, 'Data not found.', extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')
                except Exception as e:
                    messages.error(request ,f'{e}', extra_tags="alert-danger") 
                    
                    return redirect('/report/dashboardChart/')       
                                
            elif doc_type == "Delivery_Note":  
                data_points = []
                try:
                    all_fields = [f.verbose_name for f in Delivery_Note._meta.get_fields() if not f.name =="Bill_Date" if not f.name=="Invoice_No" if not f.related_model if not isinstance(f, models.FloatField) if not isinstance(f, models.BooleanField)]
                    dl_qs = Delivery_Note.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('delivery_note_details_set')
                    for dl in dl_qs:
                        d = dl.delivery_note_details_set.all()
                        total_note = d.aggregate(total_sales=Sum('Amount'))['total_sales']
                        data_points.append({
                            'label':dl.Customers.Customer_Name,
                            'y':str(total_note)
                        })
             
                    form=dashboardchartForm()     
                    return render(request, 'reportsapp/dashboard/partial_delivery_note_datatable.html',{'sales':dl_qs, 'form':form, 'fields':all_fields, 'chart':chart_type,'data_points':data_points})                
                        
                except Exception as e:
                    messages.error(request, f'{e}',extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/')
                
            elif doc_type == "salesInvoice":
                data_points=[]
                try:
                    all_fields=[f.verbose_name for f in salesInvoice._meta.get_fields() if not f .related_model if not isinstance(f, models.BooleanField)]
                    si_qs = salesInvoice.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date) 
                    if len(si_qs):
                        ch_data = [d for f in list(si_qs) for d in saleinvoicelineItem.objects.filter(parent_id=f.id)]
                        for i ,j in zip(si_qs, ch_data):
                            data_points.append({
                                'label':i.Invoice_No,
                                'y':int(j.Amount)
                            }) 
                        return render(request, 'reportsapp/dashboard/partial_sale_invoice_datatable.html', {'sales':si_qs, 'form':form, 'fields': all_fields, 'chart':chart_type, 'data_points':data_points})       
                    else:
                        messages.warning(request, 'Data not found.',extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')
                except Exception as e:
                    messages.error(request, f'e',extra_tags="alert-danger") 
                    return redirect('/report/dashboardChart/')   
                
            elif doc_type == "SalesReturn":
                dp=[]
                try:
                    all_fields = [f.verbose_name for f in SalesReturn._meta.get_fields() if not f.related_model if not isinstance(f, models.BooleanField)]
                    sr_qs = SalesReturn.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date) 
                    if len(sr_qs)>0:
                        sr_child = [d for f in list(sr_qs) for d in Sales_Return_Detail.objects.filter(entity_id=f.id)]
                        for i,j in zip(sr_qs, sr_child):
                            dp.append({
                                'label':i.Customers.Customer_Name,
                                'y':int(j.Total_Amount)
                            })    
                        return render(request, 'reportsapp/dashboard/partial_sale_return_datatable.html', {'sales':sr_qs, 'fields':all_fields, 'form':form, 'chart':chart_type, 'data_points':dp})
                    else:
                        messages.warning(request, 'Data not found.',extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')    

                except Exception as e:
                    messages.error(request, f'e',extra_tags="alert-warning")
                    return redirect('/report/dashboardChart/')    
                
                         
            elif doc_type == "PurchaseDocument":
                data_points=[]
                try:
                    all_fields=[f.verbose_name for f in PurchaseDocument._meta.get_fields() if not f.name=="PO_No" if not f.related_model if not isinstance(f, models.BooleanField)]
                    po_qs = PurchaseDocument.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('detailofpurchase_set')
                    if len(po_qs)>0:
                        for dl in po_qs:
                            d = dl.detailofpurchase_set.all()
                            total_po = d.aggregate(total_sales=Sum('Total_Amount'))['total_sales']
                            print(f"Customer: {dl.Distributors.Distributor_Name}")
                            print(f"Total Sales: {total_po}")
                            print("Sales Details:")
                            data_points.append({
                                'label':dl.Distributors.Distributor_Name,
                                'y':str(total_po)   
                            })
                        form=dashboardchartForm() 
                        return render(request, 'reportsapp/dashboard/partial_purchase_document_datatable.html',{'sales':po_qs, 'form':form, 'fields':all_fields, 'data_points':data_points, 'chart':chart_type})
                    else:
                        messages.warning(request, 'Data not found',extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')
                
                except Exception as e:
                    messages.error(request,f'e',extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/')
                
            elif doc_type == "Receipt_Note":
                data_points=[]
                try:
                    all_fields = [f.verbose_name for f in Receipt_Note._meta.get_fields() if not f.related_model if not isinstance(f, models.BooleanField)]
                    re_qs = Receipt_Note.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('receipt_note_detail_set')
                    if len(re_qs)>0:
                        for note in re_qs:
                            note_obj = note.receipt_note_detail_set.all()
                            total_note_amount = note_obj.aggregate(total_sales=Sum('Total_Amount'))['total_sales']
                            data_points.append({
                                'label':note.Distributors.Distributor_Name,
                                'y':str(total_note_amount)
                            })
                            
                        return render(request, 'reportsapp/dashboard/partial_receipt_note_datatable.html', {'sales':re_qs, 'form':form, 'chart':chart_type, 'fields':all_fields, 'data_points': data_points})
                        return redirect('/report/dashboardChart/') 
                    
                    else:
                        messages.warning(request, 'Data not found.',extra_tags="alert-warning")
                except Exception as e:
                    messages.error(request, f'e',extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/')        
                        
            elif doc_type == "PurchaseInvoice":
                data_points=[]
                try:
                    all_fields=[f.verbose_name for f in PurchaseInvoice._meta.get_fields() if not f.related_model if not isinstance(f, models.BooleanField)]
                    pi_qs = PurchaseInvoice.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('purchaseinvoicelineitem_set')
                    if len(pi_qs)>0:
                        for data in pi_qs:
                            obj = data.purchaseinvoicelineitem_set.all()
                            total_sum = obj.aggregate(total_sales=Sum('Total_Amount'))['total_sales']
                            data_points.append({
                                'label': data.Distributors.Distributor_Name,
                                'y':str(total_sum)
                            })
                        form=dashboardchartForm()
                        return render(request, 'reportsapp/dashboard/partial_purchase_invoice_datatable.html', {'sales':pi_qs, 'form':form, 'chart':chart_type, 'fields':all_fields, 'data_points':data_points})
                    else:
                        messages.warning(request, 'Data not found.',extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')
                except Exception as e:
                    messages.error(request, f'e',extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/') 
                
                
            elif doc_type == "PurchaseReturnDocument":
                data_points = []
                fields_names = [field.verbose_name for field in PurchaseReturnDocument._meta.get_fields() if not field.name == "isactive" if not field.related_model]
                try:
                    queryset = PurchaseReturnDocument.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('detailsofpurchasereturn_set') 
                    if len(queryset)>0:
                        for data in queryset:
                            obj = data.detailsofpurchasereturn_set.all()
                            total_sum = obj.aggregate(total_sales=Sum('Total_Amount'))['total_sales']
                            data_points.append({
                                'label': data.Distributors.Distributor_Name,
                                'y':str(total_sum)
                            })
                        form=dashboardchartForm()
                        return render(request, 'reportsapp/dashboard/partial_purchase_return_datatable.html', {'sales':queryset, 'form':form, 'chart':chart_type, 'fields':fields_names, 'data_points':data_points})
                        
                    messages.warning(request, "Data not Found.",extra_tags="alert-warning")
                    return redirect('/report/dashboardChart/')
                
                except Exception as e:
                    messages.error(request, f'{e}', extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/')
                    
                
            elif  doc_type =="Opening Balance":
                data_points = []
                field_names = [field.verbose_name for field in warehouseStock._meta.get_fields() if not field.name =="isactive" if not field.related_model]
                try:
                    queryset=warehouseStock.objects.filter(created_on__date__lte=to_date, created_on__date__gte=from_date).prefetch_related('warehouseitemchild_set')
                    if len(queryset)>0:
                        for data in queryset:
                            obj = data.warehouseitemchild_set.all()
                            total_sum = obj.aggregate(total_sum=Sum('Total_Amount'))['total_sum']
                            data_points.append({
                                'label':data.warehouseName,
                                'y':str(total_sum),
                            })
                            
                        form=dashboardchartForm()    
                        return render(request, 'reportsapp/dashboard/partial_opening_balance_datatable.html', {'sales':queryset, 'form':form, 'chart':chart_type, 'fields':field_names, 'data_points':data_points})

                    else:
                        messages.warning(request, 'Data not found.',extra_tags="alert-warning")
                        return redirect('/report/dashboardChart/')
                except Exception as e:
                    messages.error(request, f'Error Code: {e}',extra_tags="alert-danger")
                    return redirect('/report/dashboardChart/') 
                    
                    
        else:
            messages.warning(request, 'Please select Document Type.',extra_tags="alert-warning")  
            return redirect('/report/dashboardChart/')                     
        
    context = {
        'form':form
    }
    return render(request, 'reportsapp/dashboard/dashboardbase.html', context)            




from django.contrib.auth import get_user_model

UserModel = get_user_model()






from account.models import stockSummary

def stocksummaryReport(request):
    form = warehouseclosingBalance()
    if request.method == "POST":
        w_name = request.POST.get('w_name')
        selected_items = request.POST.getlist('groups[]')
        date = request.POST.get('date')
        dm  = [d for f in list(selected_items) for d in stockSummary.objects.filter(item=int(f), warehouse=w_name)]     
        return render(request, 'reportsapp/stockSummary/stocksummary_datatable.html', {'stocks':dm, 'form':form})    

    form = warehouseclosingBalance()
    context = {
        'form':form,
        'stocks': stockSummary.objects.none()
    }
    return render(request, 'reportsapp/stockSummary/stocksummary_datatable.html', context)
    